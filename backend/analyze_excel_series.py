#!/usr/bin/env python3
"""
详细分析Excel文件的系列分布问题
"""
import openpyxl
from openpyxl.utils import get_column_letter

def parse_merged_cells(ws):
    """解析合并单元格"""
    merged_info = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
        value = ws.cell(row=min_row, column=min_col).value
        merged_info[(min_row, min_col)] = {
            'value': value,
            'max_col': max_col,
            'max_row': max_row
        }
    return merged_info

def analyze_excel(filepath):
    """分析Excel结构"""
    print(f"分析文件: {filepath}")
    print("=" * 80)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    print(f"工作表: {ws.title}")
    print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")

    merged_info = parse_merged_cells(ws)

    # 解析系列（第1行，从F列开始）
    print("\n【系列范围映射】")
    series_ranges = {}  # name -> list of (start, end)
    processed_cols = set()

    for col in range(6, ws.max_column + 1):
        if col in processed_cols:
            continue
        if (1, col) in merged_info:
            info = merged_info[(1, col)]
            series_name = str(info['value']).strip() if info['value'] else None
            if series_name:
                for c in range(col, info['max_col'] + 1):
                    processed_cols.add(c)
                if series_name not in series_ranges:
                    series_ranges[series_name] = []
                series_ranges[series_name].append((col, info['max_col']))

    for series_name, ranges in series_ranges.items():
        print(f"\n{series_name}:")
        for start, end in ranges:
            print(f"  列范围: {get_column_letter(start)}{start}-{get_column_letter(end)}{end}")

    # 检查系列是否交错
    print("\n" + "=" * 80)
    print("【系列交错检查】")
    all_ranges = []
    for series_name, ranges in series_ranges.items():
        for start, end in ranges:
            all_ranges.append((start, end, series_name))

    # 按起始列排序
    all_ranges.sort()

    print("\n按列顺序排列的系列分布:")
    prev_series = None
    prev_end = 0
    for start, end, series_name in all_ranges:
        if prev_series and prev_series != series_name:
            print(f"  ⚠️ 列{prev_end+1}处系列切换: {prev_series} -> {series_name}")
        print(f"  列{start}-{end}: {series_name}")
        prev_series = series_name
        prev_end = end

    # 详细检查每个型号
    print("\n" + "=" * 80)
    print("【型号详细分布】")

    # 构建完整的系列映射（每个列属于哪个系列）
    col_to_series = {}
    for series_name, ranges in series_ranges.items():
        for start, end in ranges:
            for c in range(start, end + 1):
                col_to_series[c] = series_name

    # 解析型号
    for col in range(6, ws.max_column + 1):
        if (2, col) in merged_info:
            info = merged_info[(2, col)]
            model_name = str(info['value']).split('//')[0].strip() if info['value'] else None
            model_end_col = info['max_col']
        else:
            cell_value = ws.cell(row=2, column=col).value
            model_name = str(cell_value).split('//')[0].strip() if cell_value else None
            model_end_col = col + 3

        if model_name:
            # 获取该型号所属系列
            series_name = col_to_series.get(col, "未知")
            print(f"  列{get_column_letter(col)}-{get_column_letter(model_end_col)}: {model_name:30s} (系列: {series_name})")

if __name__ == "__main__":
    filepath = "/Users/xiami/Downloads/Spec/Export_SpecExcel_20260529102338.xlsx"
    analyze_excel(filepath)
