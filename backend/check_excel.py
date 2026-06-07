#!/usr/bin/env python3
"""
检查Excel文件中的系列和型号分布
"""
import openpyxl
import sys

def parse_merged_cells(ws):
    """解析合并单元格"""
    merged_info = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
        value = ws.cell(row=min_row, column=min_col).value
        merged_info[(min_row, min_col)] = {
            'value': value,
            'max_col': max_col,
            'max_row': max_row
        }
    return merged_info

def analyze_excel(filepath):
    """分析Excel结构"""
    print(f"分析文件: {filepath}")
    print("=" * 80)

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        print(f"打开文件失败: {e}")
        return

    print(f"工作表: {ws.title}")
    print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")

    merged_info = parse_merged_cells(ws)

    # 解析系列（第1行，从F列开始）
    print("\n【系列分布 - 第1行】")
    series_map = {}
    processed_cols = set()

    for col in range(6, ws.max_column + 1):
        if col in processed_cols:
            continue
        if (1, col) in merged_info:
            info = merged_info[(1, col)]
            series_name = str(info['value']).strip() if info['value'] else None
            if series_name:
                for c in range(col, info['max_col'] + 1):
                    processed_cols.add(c)
                series_map[col] = series_name
                print(f"  列{col}-{info['max_col']}: {series_name}")

    # 解析型号（第2行）
    print("\n【型号分布 - 第2行】")
    for col in range(6, ws.max_column + 1):
        if (2, col) in merged_info:
            info = merged_info[(2, col)]
            model_name = str(info['value']).split('//')[0].strip() if info['value'] else None
            if model_name:
                # 查找该列属于哪个系列
                series_name = "未知"
                for s_col, s_name in series_map.items():
                    if s_col <= col <= s_col + 100:  # 粗略匹配
                        series_name = s_name
                        break
                print(f"  列{col}-{info['max_col']}: {model_name} (所属系列: {series_name})")

    # 检查是否有 VINNO 6 综合版
    print("\n【查找 VINNO 6 综合版】")
    found_models = []
    for col in range(6, ws.max_column + 1):
        if (2, col) in merged_info:
            info = merged_info[(2, col)]
            model_name = str(info['value']).split('//')[0].strip() if info['value'] else None
            if model_name and "综合版" in model_name:
                # 查找所属系列
                series_name = "未知"
                for s_col in sorted(series_map.keys()):
                    s_end = s_col + 100  # 估算
                    if s_col <= col <= s_end:
                        series_name = series_map[s_col]
                        break
                found_models.append((model_name, series_name, col))
                print(f"  ✅ 找到: {model_name} 在列{col} (系列: {series_name})")

    if not found_models:
        print("  ❌ 未找到任何包含'综合版'的型号")

if __name__ == "__main__":
    filepath = "/Users/xiami/Downloads/Spec/Export_SpecExcel_20260529102338.xlsx"
    analyze_excel(filepath)
