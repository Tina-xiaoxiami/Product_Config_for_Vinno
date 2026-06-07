#!/usr/bin/env python3
"""
检查多个Excel文件的系列分布
"""
import openpyxl
from openpyxl.utils import get_column_letter
import os

def parse_merged_cells(ws):
    """解析合并单元格"""
    merged_info = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
        value = ws.cell(row=min_row, column=min_col).value
        merged_info[(min_row, min_col)] = {
            'value': value,
            'max_col': max_col,
            'max_row': max_row
        }
    return merged_info

def analyze_excel(filepath):
    """分析Excel结构"""
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        return

    print(f"\n分析文件: {os.path.basename(filepath)}")
    print("=" * 80)

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        print(f"打开文件失败: {e}")
        return

    print(f"工作表: {ws.title}, 行: {ws.max_row}, 列: {ws.max_column}")

    merged_info = parse_merged_cells(ws)

    # 解析系列（第1行，从F列开始）
    print("\n【系列分布】")
    series_ranges = []
    processed_cols = set()

    for col in range(6, ws.max_column + 1):
        if col in processed_cols:
            continue
        if (1, col) in merged_info:
            info = merged_info[(1, col)]
            series_name = str(info['value']).strip() if info['value'] else None
            if series_name:
                for c in range(col, info['max_col'] + 1):
                    processed_cols.add(c)
                series_ranges.append((series_name, col, info['max_col']))
                print(f"  {series_name}: 列{get_column_letter(col)}-{get_column_letter(info['max_col'])} ({col}-{info['max_col']})")

    # 解析型号（第2行）
    print("\n【型号分布】")
    col_to_series = {}
    for name, start, end in series_ranges:
        for c in range(start, end + 1):
            col_to_series[c] = name

    for col in range(6, ws.max_column + 1):
        if (2, col) in merged_info:
            info = merged_info[(2, col)]
            model_name = str(info['value']).split('//')[0].strip() if info['value'] else None
            if model_name:
                series_name = col_to_series.get(col, "未知")
                print(f"  列{get_column_letter(col)}: {model_name:40s} (系列: {series_name})")

if __name__ == "__main__":
    files = [
        "/Users/xiami/Downloads/Spec/Export_SpecExcel_20260529102338.xlsx",
        "/Users/xiami/Downloads/Spec/Export_SpecExcel_20260529102149.xlsx",
        "/Users/xiami/Downloads/Spec/Export_SpecExcel_20260529101436.xlsx"
    ]

    for filepath in files:
        analyze_excel(filepath)
