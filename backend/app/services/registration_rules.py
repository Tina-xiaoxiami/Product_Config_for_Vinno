"""国内注册差异表解析与注册红线判定规则。"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook


_EMPTY_VALUES = {"", "-", "N/A", "未定义", "NONE", "NULL"}
_FORMAL_STRATEGY_VALUES = {"X", "O", "Δ"}


@dataclass(frozen=True)
class RegistrationProbeSource:
    model: str
    ipn: str
    source_row: int


@dataclass(frozen=True)
class RegistrationModelSource:
    model_name: str
    unsupported_probes: tuple[str, ...]
    channel_count: int | None
    source_row: int


@dataclass(frozen=True)
class DomesticRegistrationWorkbook:
    probes: tuple[RegistrationProbeSource, ...]
    models: tuple[RegistrationModelSource, ...]


@dataclass(frozen=True)
class ProbeAvailability:
    effective_status: str
    status_source: str
    is_formal: bool
    conflict: bool = False


def normalize_business_name(value: object) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or ""))
    return re.sub(r"\s+", " ", normalized).strip()


def _split_probe_names(value: object) -> tuple[str, ...]:
    text = normalize_business_name(value)
    if not text or "探头全适用" in text:
        return ()
    return tuple(
        part.strip()
        for part in re.split(r"[，、,;\n]+", text)
        if part.strip()
    )


def _find_model_header_row(sheet) -> int:
    for row in range(1, min(sheet.max_row, 20) + 1):
        values = {
            normalize_business_name(sheet.cell(row=row, column=column).value)
            for column in range(1, min(sheet.max_column, 12) + 1)
        }
        if {"型号", "不支持探头"} <= values:
            return row
    raise ValueError("注册差异表缺少‘型号/不支持探头’表头")


def _ensure_sheet_dimensions(sheet) -> None:
    """让未携带 worksheet dimension 元数据的合法工作簿可被流式解析。"""

    if sheet.max_row is None or sheet.max_column is None:
        sheet.calculate_dimension(force=True)


def parse_domestic_registration_workbook(
    workbook_path: str | Path,
) -> DomesticRegistrationWorkbook:
    """解析已确认的国内型号差异表，不修改原文件。"""

    path = Path(workbook_path)
    if not path.is_file():
        raise FileNotFoundError(path)

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        matrix_sheet = workbook["0729"] if "0729" in workbook.sheetnames else workbook.worksheets[0]
        probe_sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.worksheets[1]
        _ensure_sheet_dimensions(matrix_sheet)
        _ensure_sheet_dimensions(probe_sheet)

        probes: list[RegistrationProbeSource] = []
        seen_probe_models: set[str] = set()
        seen_probe_ipns: set[str] = set()
        for row in range(1, probe_sheet.max_row + 1):
            model = normalize_business_name(probe_sheet.cell(row=row, column=1).value)
            ipn = normalize_business_name(probe_sheet.cell(row=row, column=2).value)
            if not model and not ipn:
                continue
            if not model or not ipn:
                raise ValueError(f"探头IPN表第 {row} 行不完整")
            if model in seen_probe_models:
                raise ValueError(f"探头型号重复：{model}")
            if ipn in seen_probe_ipns:
                raise ValueError(f"探头IPN重复：{ipn}")
            seen_probe_models.add(model)
            seen_probe_ipns.add(ipn)
            probes.append(RegistrationProbeSource(model=model, ipn=ipn, source_row=row))

        declared_probes = set(_split_probe_names(matrix_sheet.cell(row=2, column=2).value))
        if declared_probes and declared_probes != seen_probe_models:
            missing = sorted(declared_probes - seen_probe_models)
            extra = sorted(seen_probe_models - declared_probes)
            raise ValueError(f"支持探头清单与IPN表不一致：缺少={missing}，多出={extra}")

        header_row = _find_model_header_row(matrix_sheet)
        headers = {
            normalize_business_name(matrix_sheet.cell(row=header_row, column=column).value): column
            for column in range(1, matrix_sheet.max_column + 1)
        }
        model_column = headers["型号"]
        unsupported_column = headers["不支持探头"]
        channel_column = headers.get("通道数")
        models: list[RegistrationModelSource] = []
        seen_models: set[str] = set()
        for row in range(header_row + 1, matrix_sheet.max_row + 1):
            model_name = normalize_business_name(
                matrix_sheet.cell(row=row, column=model_column).value
            )
            if not model_name:
                continue
            if model_name in seen_models:
                raise ValueError(f"注册型号重复：{model_name}")
            unsupported = _split_probe_names(
                matrix_sheet.cell(row=row, column=unsupported_column).value
            )
            unknown = sorted(set(unsupported) - seen_probe_models)
            if unknown:
                raise ValueError(
                    f"型号 {model_name} 引用未知不支持探头：{', '.join(unknown)}"
                )
            channel_value = (
                matrix_sheet.cell(row=row, column=channel_column).value
                if channel_column
                else None
            )
            channel_count = int(channel_value) if channel_value not in (None, "") else None
            seen_models.add(model_name)
            models.append(
                RegistrationModelSource(
                    model_name=model_name,
                    unsupported_probes=unsupported,
                    channel_count=channel_count,
                    source_row=row,
                )
            )

        if not probes or not models:
            raise ValueError("注册差异表未解析到型号或探头")
        return DomesticRegistrationWorkbook(probes=tuple(probes), models=tuple(models))
    finally:
        workbook.close()


def _normalize_strategy_value(value: object) -> str:
    normalized = normalize_business_name(value).upper().replace("∆", "Δ")
    return "" if normalized in _EMPTY_VALUES else normalized


def evaluate_probe_availability(
    *,
    registered: bool,
    selection_config: object,
    current_config: object,
) -> ProbeAvailability:
    """先执行注册红线，再使用正式选型，最后才参考研发当前配置。"""

    selection = _normalize_strategy_value(selection_config)
    current = _normalize_strategy_value(current_config)
    if not registered:
        return ProbeAvailability(
            effective_status="#",
            status_source="registration_redline",
            is_formal=True,
            conflict=selection in _FORMAL_STRATEGY_VALUES,
        )
    if selection in _FORMAL_STRATEGY_VALUES:
        return ProbeAvailability(
            effective_status=selection,
            status_source="selection_config",
            is_formal=True,
        )
    if current in _FORMAL_STRATEGY_VALUES:
        return ProbeAvailability(
            effective_status=current,
            status_source="current_config_aux",
            is_formal=False,
        )
    return ProbeAvailability(
        effective_status="未定义",
        status_source="missing",
        is_formal=False,
        conflict=selection == "#",
    )
