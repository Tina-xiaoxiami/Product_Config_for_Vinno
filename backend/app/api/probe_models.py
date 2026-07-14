"""探头型号 API"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete
import io
from datetime import datetime
from urllib.parse import quote
from app.database import get_db
from app.models.probe import ProbeModel, ProbeModelVariant, ProbeModelApp, Application, ProbeCategory
from app.models.config import ConfigItem
from app.schemas.probe import (
    ProbeModelCreate, ProbeModelUpdate,
    ProbeModelResponse, ProbeModelListResponse,
)

router = APIRouter()


@router.get("", response_model=ProbeModelListResponse)
async def list_models(
    category_id: int = None,
    skip: int = 0, limit: int = 200,
    db: AsyncSession = Depends(get_db)
):
    q = select(ProbeModel)
    cq = select(func.count()).select_from(ProbeModel)
    if category_id:
        q = q.where(ProbeModel.category_id == category_id)
        cq = cq.where(ProbeModel.category_id == category_id)
    count_r = await db.execute(cq)
    total = count_r.scalar()
    result = await db.execute(q.order_by(ProbeModel.sort_order).offset(skip).limit(limit))
    items = result.scalars().all()
    return ProbeModelListResponse(items=items, total=total)


@router.post("", response_model=ProbeModelResponse)
async def create_model(data: ProbeModelCreate, db: AsyncSession = Depends(get_db)):
    obj = ProbeModel(**data.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


@router.put("/{model_id}", response_model=ProbeModelResponse)
async def update_model(model_id: int, data: ProbeModelUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(ProbeModel).where(ProbeModel.id == model_id))
    obj = result.scalar_one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="探头型号不存在")
    for f, v in data.model_dump(exclude_unset=True).items():
        setattr(obj, f, v)
    await db.commit()
    await db.refresh(obj)
    return obj


@router.delete("/{model_id}")
async def delete_model(model_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(ProbeModel).where(ProbeModel.id == model_id))
    obj = result.scalar_one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="探头型号不存在")
    await db.delete(obj)
    await db.commit()
    return {"message": "删除成功"}


# ===== Probe Model Applications =====

@router.get("/{model_id}/apps")
async def get_model_apps(model_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(ProbeModelApp, Application)
        .join(Application, ProbeModelApp.application_id == Application.id)
        .where(ProbeModelApp.probe_model_id == model_id)
        .order_by(Application.sort_order)
    )
    return [{"id": a.Application.id, "name": a.Application.name} for _, a in result.all()]


@router.post("/{model_id}/apps")
async def set_model_apps(model_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    """设置探头型号支持的应用列表 { app_ids: [1,2,3] }"""
    # Delete existing
    await db.execute(delete(ProbeModelApp).where(ProbeModelApp.probe_model_id == model_id))
    # Add new
    for aid in data.get("app_ids", []):
        db.add(ProbeModelApp(probe_model_id=model_id, application_id=aid))
    await db.commit()
    return {"message": f"已设置 {len(data.get('app_ids', []))} 个应用"}


# ===== Variant Excel Export/Import =====

@router.get("/variants/export")
async def export_variants_excel(db: AsyncSession = Depends(get_db)):
    """导出所有探头型号变体为Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    result = await db.execute(
        select(ProbeModelVariant, ProbeModel, ProbeCategory)
        .join(ProbeModel, ProbeModelVariant.probe_model_id == ProbeModel.id)
        .join(ProbeCategory, ProbeModel.category_id == ProbeCategory.id)
        .order_by(ProbeCategory.sort_order, ProbeModel.sort_order, ProbeModelVariant.sort_order)
    )
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "探头型号变体"

    header_font = Font(bold=True, color="ffffff")
    header_fill = PatternFill("solid", fgColor="409EFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["探头类别", "对外型号", "内部型号", "IPN", "备注", "排序"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, (v, m, cat) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=cat.name).border = thin_border
        ws.cell(row=i, column=2, value=m.model_number).border = thin_border
        ws.cell(row=i, column=3, value=v.internal_model).border = thin_border
        ws.cell(row=i, column=4, value=v.ipn or "").border = thin_border
        ws.cell(row=i, column=5, value=v.notes or "").border = thin_border
        ws.cell(row=i, column=6, value=v.sort_order or 0).border = thin_border

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 8

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"探头型号变体_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


@router.post("/variants/import")
async def import_variants_excel(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """从Excel导入探头型号变体"""
    from openpyxl import load_workbook

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 文件")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="Excel文件为空")

    # Build model lookup: model_number -> ProbeModel
    model_result = await db.execute(select(ProbeModel))
    models = {m.model_number: m for m in model_result.scalars().all()}

    added = 0; updated = 0; skipped = 0; errors = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1] or not row[2]:
            skipped += 1; continue
        cat_name = str(row[0]).strip()
        model_number = str(row[1]).strip()
        internal_model = str(row[2]).strip()
        ipn = str(row[3]).strip() if row[3] else None
        notes = str(row[4]).strip() if len(row) > 4 and row[4] else None
        sort_order = int(row[5]) if len(row) > 5 and row[5] is not None else 0

        if not internal_model:
            skipped += 1; continue

        probe_model = models.get(model_number)
        if not probe_model:
            errors.append(f"第{row[0] if hasattr(row, 'row') else ''}行: 未找到对外型号 '{model_number}'")
            skipped += 1; continue

        # Check if variant already exists for this model
        vr = await db.execute(
            select(ProbeModelVariant).where(
                ProbeModelVariant.probe_model_id == probe_model.id,
                ProbeModelVariant.internal_model == internal_model
            )
        )
        existing = vr.scalar_one_or_none()
        if existing:
            existing.ipn = ipn
            existing.notes = notes
            existing.sort_order = sort_order
            updated += 1
        else:
            db.add(ProbeModelVariant(
                probe_model_id=probe_model.id,
                internal_model=internal_model,
                ipn=ipn,
                notes=notes,
                sort_order=sort_order
            ))
            added += 1

    await db.commit()
    return {
        "message": f"导入完成: 新增 {added} 条, 更新 {updated} 条, 跳过 {skipped} 条",
        "added": added, "updated": updated, "skipped": skipped,
        "errors": errors[:10]  # Return first 10 errors
    }


@router.get("/{model_id}/variants/export")
async def export_model_variants_excel(model_id: int, db: AsyncSession = Depends(get_db)):
    """导出单个探头型号的变体为Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    r = await db.execute(select(ProbeModel).where(ProbeModel.id == model_id))
    model = r.scalar_one_or_none()
    if not model:
        raise HTTPException(status_code=404, detail="探头型号不存在")

    cat_r = await db.execute(select(ProbeCategory).where(ProbeCategory.id == model.category_id))
    cat = cat_r.scalar_one_or_none()

    vr = await db.execute(
        select(ProbeModelVariant).where(ProbeModelVariant.probe_model_id == model_id)
        .order_by(ProbeModelVariant.sort_order)
    )
    variants = vr.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "变体"

    header_font = Font(bold=True, color="ffffff")
    header_fill = PatternFill("solid", fgColor="409EFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["探头类别", "对外型号", "内部型号", "IPN", "备注", "排序"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, v in enumerate(variants, 2):
        ws.cell(row=i, column=1, value=cat.name if cat else "").border = thin_border
        ws.cell(row=i, column=2, value=model.model_number).border = thin_border
        ws.cell(row=i, column=3, value=v.internal_model).border = thin_border
        ws.cell(row=i, column=4, value=v.ipn or "").border = thin_border
        ws.cell(row=i, column=5, value=v.notes or "").border = thin_border
        ws.cell(row=i, column=6, value=v.sort_order or 0).border = thin_border

    # Empty rows for adding new variants
    for i in range(len(variants) + 2, len(variants) + 12):
        ws.cell(row=i, column=1, value=cat.name if cat else "").border = thin_border
        ws.cell(row=i, column=2, value=model.model_number).border = thin_border
        for c in range(3, 7):
            ws.cell(row=i, column=c).border = thin_border

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 8

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{model.model_number}_变体.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


@router.post("/{model_id}/variants/import")
async def import_model_variants_excel(model_id: int, file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """从Excel导入单个探头型号的变体"""
    from openpyxl import load_workbook

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 文件")

    r = await db.execute(select(ProbeModel).where(ProbeModel.id == model_id))
    model = r.scalar_one_or_none()
    if not model:
        raise HTTPException(status_code=404, detail="探头型号不存在")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="Excel文件为空")

    added = 0; updated = 0; skipped = 0; errors = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3: skipped += 1; continue
        internal_model = str(row[2]).strip() if row[2] else ""
        if not internal_model: skipped += 1; continue
        ipn = str(row[3]).strip() if len(row) > 3 and row[3] else None
        notes = str(row[4]).strip() if len(row) > 4 and row[4] else None
        sort_order = int(row[5]) if len(row) > 5 and row[5] is not None else 0

        vr = await db.execute(
            select(ProbeModelVariant).where(
                ProbeModelVariant.probe_model_id == model_id,
                ProbeModelVariant.internal_model == internal_model
            )
        )
        existing = vr.scalar_one_or_none()
        if existing:
            existing.ipn = ipn
            existing.notes = notes
            existing.sort_order = sort_order
            updated += 1
        else:
            db.add(ProbeModelVariant(probe_model_id=model_id, internal_model=internal_model, ipn=ipn, notes=notes, sort_order=sort_order))
            added += 1

    await db.commit()
    return {
        "message": f"导入完成: 新增 {added} 条, 更新 {updated} 条, 跳过 {skipped} 条",
        "added": added, "updated": updated, "skipped": skipped,
        "errors": errors[:10]
    }


@router.post("/variants/auto-populate")
async def auto_populate_variants(db: AsyncSession = Depends(get_db)):
    """从 config_items（Probes 类别）自动填充探头型号变体

    映射关系：
      config_items.en_desc → probe_models.model_number（型号=英文名称）
      config_items.rd_name → ProbeModelVariant.internal_model（内部型号=研发名称）
      config_items.ipn     → ProbeModelVariant.ipn（IPN=IPN）
    """
    # 读取 config_items 中的 Probes 数据
    ci_result = await db.execute(
        select(ConfigItem).where(ConfigItem.category == "Probes")
    )
    config_items = ci_result.scalars().all()
    if not config_items:
        raise HTTPException(status_code=404, detail="config_items 中没有 Probes 类别的数据")

    # 读取所有探头型号
    pm_result = await db.execute(select(ProbeModel))
    probe_models = pm_result.scalars().all()

    # 构建 model_number → ProbeModel 的索引（小写，去括号/后缀）
    def normalize_key(en_desc):
        if not en_desc: return ""
        key = str(en_desc).strip().lower()
        # 去掉括号内的说明
        for sep in ['(', '（', ',']:
            key = key.split(sep)[0].strip()
        # 去掉特殊后缀
        for suffix in [' compatible', ' ngs']:
            key = key.replace(suffix, '').strip()
        return key

    pm_map = {}
    for pm in probe_models:
        key = pm.model_number.strip().lower()
        if key not in pm_map:
            pm_map[key] = []
        pm_map[key].append(pm)

    added = 0
    skipped_no_match = 0
    skipped_dup = 0
    matched_models = set()

    for ci in config_items:
        if not ci.en_desc:
            skipped_no_match += 1
            continue

        key = normalize_key(ci.en_desc)
        candidates = pm_map.get(key, [])
        if not candidates:
            skipped_no_match += 1
            continue

        for pm in candidates:
            matched_models.add(pm.id)
            # 检查是否已有相同 internal_model 的变体
            vr = await db.execute(
                select(ProbeModelVariant).where(
                    ProbeModelVariant.probe_model_id == pm.id,
                    ProbeModelVariant.internal_model == ci.rd_name
                )
            )
            if vr.scalar_one_or_none():
                skipped_dup += 1
                continue

            # 创建变体
            internal = ci.rd_name.strip() if ci.rd_name else ""
            ipn = ci.ipn.strip() if ci.ipn else None
            if not internal:
                skipped_no_match += 1
                continue

            # 去掉 rd_name 中的 【启用】/【停用】标记
            clean_internal = internal.replace("【启用】", "").replace("【停用】", "").strip()

            db.add(ProbeModelVariant(
                probe_model_id=pm.id,
                internal_model=clean_internal,
                ipn=ipn,
                sort_order=0
            ))
            added += 1

    await db.commit()
    return {
        "message": f"自动填充完成：新增 {added} 条变体，跳过 {skipped_dup} 条重复，{skipped_no_match} 条未匹配",
        "added": added,
        "skipped_duplicate": skipped_dup,
        "skipped_no_match": skipped_no_match,
        "matched_models": len(matched_models),
    }


# ===== Probe Model Variants =====

@router.get("/{model_id}/variants")
async def get_variants(model_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(ProbeModelVariant).where(ProbeModelVariant.probe_model_id == model_id).order_by(ProbeModelVariant.sort_order)
    )
    return [{"id": v.id, "internal_model": v.internal_model, "ipn": v.ipn, "notes": v.notes or "", "sort_order": v.sort_order} for v in result.scalars().all()]


@router.post("/{model_id}/variants")
async def save_variant(model_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    """保存变体 { internal_model, ipn, notes, sort_order }。有id则更新，无id则新增"""
    vid = data.get("id")
    internal = data.get("internal_model", "").strip()
    ipn = data.get("ipn", "").strip() or None
    notes = data.get("notes", "").strip() or None
    sort = data.get("sort_order", 0)
    if not internal: raise HTTPException(status_code=400, detail="内部型号不能为空")

    if vid:
        r = await db.execute(select(ProbeModelVariant).where(ProbeModelVariant.id == vid))
        v = r.scalar_one_or_none()
        if not v: raise HTTPException(status_code=404, detail="变体不存在")
        v.internal_model = internal; v.ipn = ipn; v.notes = notes; v.sort_order = sort
    else:
        v = ProbeModelVariant(probe_model_id=model_id, internal_model=internal, ipn=ipn, notes=notes, sort_order=sort)
        db.add(v)
    await db.commit(); await db.refresh(v)
    return {"message": "保存成功", "id": v.id}


@router.delete("/{model_id}/variants/{variant_id}")
async def delete_variant(model_id: int, variant_id: int, db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(ProbeModelVariant).where(ProbeModelVariant.id == variant_id))
    v = r.scalar_one_or_none()
    if not v: raise HTTPException(status_code=404, detail="变体不存在")
    await db.delete(v); await db.commit()
    return {"message": "删除成功"}
