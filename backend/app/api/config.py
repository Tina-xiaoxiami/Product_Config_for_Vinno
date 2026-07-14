"""
配置数据 API
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List, Optional
import json
import io
from datetime import datetime

from app.database import get_db
from app.models import ConfigItem, ConfigValue, ProductModel, ProductSeries
from app.schemas.config import (
    ConfigRowResponse, ConfigValueResponse,
    ConfigCompareRequest, ConfigCompareResponse, ConfigDiffItem,
    BatchUpdateRequest
)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter()


@router.get("/rows")
async def get_config_rows(
    series_id: int,
    categories: Optional[str] = None,  # 逗号分隔的分类列表
    search: Optional[str] = None,
    include_empty: bool = Query(False, description="是否包含所有配置值为空的行（草稿筛选时使用）"),
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db)
):
    """获取配置数据行

    Args:
        categories: 逗号分隔的分类列表，如 "Optional Features,Probes"
        include_empty: 包含空值行，草稿筛选（特别是"删除"筛选）时使用
    """
    # 获取该系列下的所有型号
    models_result = await db.execute(
        select(ProductModel).where(ProductModel.series_id == series_id)
    )
    models = models_result.scalars().all()
    model_ids = [m.id for m in models]

    # 解析分类列表
    category_list = [c.strip() for c in categories.split(',') if c.strip()] if categories else None

    # 构建查询
    query = select(ConfigItem)

    # 分类筛选
    if category_list:
        query = query.where(ConfigItem.category.in_(category_list))

    if search:
        query = query.where(
            (ConfigItem.rd_name.contains(search)) |
            (ConfigItem.ipn.contains(search)) |
            (ConfigItem.v_code.contains(search)) |
            (ConfigItem.zh_desc.contains(search)) |
            (ConfigItem.en_desc.contains(search))
        )

    # 获取总数（在分页前）
    count_query = select(func.count()).select_from(ConfigItem)
    if category_list:
        count_query = count_query.where(ConfigItem.category.in_(category_list))
    if search:
        count_query = count_query.where(
            (ConfigItem.rd_name.contains(search)) |
            (ConfigItem.ipn.contains(search)) |
            (ConfigItem.v_code.contains(search)) |
            (ConfigItem.zh_desc.contains(search)) |
            (ConfigItem.en_desc.contains(search))
        )
    count_result = await db.execute(count_query)
    total = count_result.scalar()

    # 分页查询
    query = query.order_by(ConfigItem.row_index).offset(skip).limit(limit)
    result = await db.execute(query)
    items = result.scalars().all()

    # 获取配置值 - 一次性查询所有（避免N+1问题）
    item_ids = [item.id for item in items]
    values_result = await db.execute(
        select(ConfigValue).where(
            ConfigValue.item_id.in_(item_ids),
            ConfigValue.model_id.in_(model_ids)
        )
    )
    values = values_result.scalars().all()

    # 构建索引：item_id -> model_id -> value
    value_map = {}
    for v in values:
        if v.item_id not in value_map:
            value_map[v.item_id] = {}
        value_map[v.item_id][v.model_id] = {
            "id": v.id,
            "current_config": v.current_config,
            "final_config": v.final_config,
            "selection_config": v.selection_config,
            "rd_status": v.rd_status
        }

    # 组装结果
    rows = []
    for item in items:
        item_values = value_map.get(item.id, {})

        # 检查是否所有配置值都为空
        has_non_empty_value = False
        for model_id in model_ids:
            if model_id in item_values:
                v = item_values[model_id]
                # 检查四个字段是否有非空值
                for field in ['final_config', 'current_config', 'selection_config', 'rd_status']:
                    field_value = v.get(field)
                    if field_value and str(field_value).strip() not in ['', 'None', 'null', 'N/A', '-']:
                        has_non_empty_value = True
                        break
                if has_non_empty_value:
                    break

        # 如果所有配置值都为空，则跳过该行（除非指定了 include_empty）
        if not has_non_empty_value and not include_empty:
            continue

        rows.append({
            "id": item.id,
            "category": item.category,
            "row_index": item.row_index,
            "rd_name": item.rd_name,
            "v_code": item.v_code,
            "ipn": item.ipn,
            "zh_desc": item.zh_desc,
            "en_desc": item.en_desc,
            "model_values": item_values
        })

    # 用 SQL 总条数（分页前），而非 len(rows)（仅当前页）
    return {"items": rows, "total": total}


@router.post("/compare", response_model=ConfigCompareResponse)
async def compare_configs(
    data: ConfigCompareRequest,
    db: AsyncSession = Depends(get_db)
):
    """配置对比"""
    if len(data.model_ids) < 2:
        raise HTTPException(status_code=400, detail="至少选择2个型号进行对比")

    # 获取型号信息
    models_result = await db.execute(
        select(ProductModel).where(ProductModel.id.in_(data.model_ids))
    )
    models = {m.id: m for m in models_result.scalars().all()}

    # 获取所有配置项
    items_result = await db.execute(
        select(ConfigItem).where(ConfigItem.category != "Main Unit").order_by(ConfigItem.row_index)
    )
    items = items_result.scalars().all()

    # 获取配置值
    values_result = await db.execute(
        select(ConfigValue).where(ConfigValue.model_id.in_(data.model_ids))
    )
    values = values_result.scalars().all()

    # 构建索引: item_id -> model_id -> value
    value_map = {}
    for v in values:
        if v.item_id not in value_map:
            value_map[v.item_id] = {}
        value_map[v.item_id][v.model_id] = v

    # 找出差异项
    diff_items = []
    for item in items:
        if item.id not in value_map:
            continue

        item_values = value_map[item.id]
        if len(item_values) < 2:
            continue

        # 对每个字段检查差异
        for field in data.compare_fields:
            field_values = {}
            for model_id in data.model_ids:
                if model_id in item_values:
                    v = getattr(item_values[model_id], field, None)
                    field_values[model_id] = v

            # 检查是否有差异
            unique_values = set(v for v in field_values.values() if v not in [None, "N/A", ""])
            has_diff = len(unique_values) > 1

            # 根据 show_only_diff 决定是否返回
            if data.show_only_diff:
                # 仅返回有差异的项
                if has_diff:
                    first_model_id = list(field_values.keys())[0]
                    model = models.get(first_model_id)
                    diff_items.append({
                        "item_id": item.id,
                        "row_index": item.row_index,
                        "rd_name": item.rd_name,
                        "ipn": item.ipn,
                        "v_code": item.v_code,
                        "zh_desc": item.zh_desc,
                        "en_desc": item.en_desc,
                        "model_id": first_model_id,
                        "model_name": model.name if model else "",
                        "field_name": field,
                        "values": field_values,
                        "has_diff": True
                    })
            else:
                # 返回所有项
                first_model_id = list(field_values.keys())[0] if field_values else None
                model = models.get(first_model_id) if first_model_id else None
                diff_items.append({
                    "item_id": item.id,
                    "row_index": item.row_index,
                    "rd_name": item.rd_name,
                    "ipn": item.ipn,
                    "v_code": item.v_code,
                    "zh_desc": item.zh_desc,
                    "en_desc": item.en_desc,
                    "model_id": first_model_id,
                    "model_name": model.name if model else "",
                    "field_name": field,
                    "values": field_values,
                    "has_diff": has_diff
                })

    # 计算差异数
    diff_count = sum(1 for item in diff_items if item.get("has_diff"))

    return {
        "items": diff_items,
        "total": len(items),
        "diff_count": diff_count
    }


@router.put("/value/{value_id}")
async def update_config_value(
    value_id: int,
    current_config: Optional[str] = None,
    final_config: Optional[str] = None,
    selection_config: Optional[str] = None,
    rd_status: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """更新配置值"""
    result = await db.execute(select(ConfigValue).where(ConfigValue.id == value_id))
    value = result.scalar_one_or_none()

    if not value:
        raise HTTPException(status_code=404, detail="配置值不存在")

    if current_config is not None:
        value.current_config = current_config
    if final_config is not None:
        value.final_config = final_config
    if selection_config is not None:
        value.selection_config = selection_config
    if rd_status is not None:
        value.rd_status = rd_status

    await db.commit()
    await db.refresh(value)
    return value


@router.post("/batch-update")
async def batch_update_config_values(
    data: BatchUpdateRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    批量更新配置值

    Args:
        data: BatchUpdateRequest
            - item_ids: 配置项ID列表
            - model_ids: 型号ID列表（可选，不传则更新所有型号）
            - field_name: 字段名（current_config/final_config/selection_config/rd_status）
            - value: 新值

    Returns:
        更新的记录数
    """
    # 验证字段名
    valid_fields = ["current_config", "final_config", "selection_config", "rd_status"]
    if data.field_name not in valid_fields:
        raise HTTPException(
            status_code=400,
            detail=f"无效的字段名，必须是: {', '.join(valid_fields)}"
        )

    # 获取所有相关的配置值
    query = select(ConfigValue).where(ConfigValue.item_id.in_(data.item_ids))

    if data.model_ids:
        query = query.where(ConfigValue.model_id.in_(data.model_ids))

    result = await db.execute(query)
    values = result.scalars().all()

    if not values:
        raise HTTPException(status_code=404, detail="未找到匹配的配置值")

    # 批量更新
    updated_count = 0
    for v in values:
        setattr(v, data.field_name, data.value)
        updated_count += 1

    await db.commit()

    return {
        "message": "批量更新成功",
        "updated_count": updated_count,
        "item_count": len(data.item_ids),
        "model_count": len(data.model_ids) if data.model_ids else "全部"
    }


@router.post("/compare/export")
async def export_compare_result(
    data: ConfigCompareRequest,
    db: AsyncSession = Depends(get_db)
):
    """导出对比结果为Excel"""
    # 获取对比数据（复用compare_configs的逻辑）
    if len(data.model_ids) < 2:
        raise HTTPException(status_code=400, detail="至少选择2个型号进行对比")

    # 获取型号信息
    models_result = await db.execute(
        select(ProductModel).where(ProductModel.id.in_(data.model_ids))
    )
    models = {m.id: m for m in models_result.scalars().all()}

    # 获取所有配置项
    items_result = await db.execute(
        select(ConfigItem).where(ConfigItem.category != "Main Unit").order_by(ConfigItem.row_index)
    )
    items = items_result.scalars().all()

    # 获取配置值
    values_result = await db.execute(
        select(ConfigValue).where(ConfigValue.model_id.in_(data.model_ids))
    )
    values = values_result.scalars().all()

    # 构建索引
    value_map = {}
    for v in values:
        if v.item_id not in value_map:
            value_map[v.item_id] = {}
        value_map[v.item_id][v.model_id] = v

    # 收集对比数据
    compare_data = []
    for item in items:
        if item.id not in value_map:
            continue

        item_values = value_map[item.id]
        if len(item_values) < 2:
            continue

        for field in data.compare_fields:
            field_values = {}
            for model_id in data.model_ids:
                if model_id in item_values:
                    v = getattr(item_values[model_id], field, None)
                    field_values[model_id] = v

            unique_values = set(v for v in field_values.values() if v not in [None, "N/A", ""])
            has_diff = len(unique_values) > 1

            if data.show_only_diff and not has_diff:
                continue

            compare_data.append({
                "rd_name": item.rd_name,
                "ipn": item.ipn,
                "v_code": item.v_code,
                "zh_desc": item.zh_desc,
                "en_desc": item.en_desc,
                "field_name": field,
                "values": field_values,
                "has_diff": has_diff
            })

    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "配置对比结果"

    # 样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    diff_fill = PatternFill(start_color="FDF6EC", end_color="FDF6EC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    field_labels = {
        "final_config": "最终配置",
        "current_config": "当前配置",
        "selection_config": "选型类别",
        "rd_status": "研发状态"
    }

    headers = ["研发名称", "中文名称", "英文名称", "V代码", "IPN号", "对比字段", "差异"] + [models[mid].name for mid in data.model_ids]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # 数据行
    for row_idx, item in enumerate(compare_data, 2):
        ws.cell(row=row_idx, column=1, value=item["rd_name"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=item.get("zh_desc", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=item.get("en_desc", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=item.get("v_code", "")).border = thin_border
        ws.cell(row=row_idx, column=5, value=item["ipn"]).border = thin_border
        ws.cell(row=row_idx, column=6, value=field_labels.get(item["field_name"], item["field_name"])).border = thin_border
        ws.cell(row=row_idx, column=7, value="是" if item["has_diff"] else "否").border = thin_border

        for col_idx, model_id in enumerate(data.model_ids, 8):
            value = item["values"].get(model_id, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value or "-")
            cell.border = thin_border
            # 差异项高亮
            if item["has_diff"]:
                cell.fill = diff_fill

    # 列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    for col_idx in range(5, 5 + len(data.model_ids)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

    # 导出
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"配置对比_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )