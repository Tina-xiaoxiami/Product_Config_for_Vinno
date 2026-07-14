"""探头配置 Excel 导入 API"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import io, json
from app.database import get_db
from app.models.probe import (
    ProbeCategory, ProbeModel, Application, CategoryApplication,
    FeatureGroup, Feature, TemplateFeature,
    ProductProbeModel, ProductProbeConfig,
)
from app.models.product_model import ProductModel
import openpyxl

router = APIRouter()

NA_VALUES = {'', None, 'N/A', '-', 'None', 'null', '未定义'}


def _parse_merged_cells(ws):
    merged = {}
    for rng in ws.merged_cells.ranges:
        merged[(rng.min_row, rng.min_col)] = {
            'value': ws.cell(row=rng.min_row, column=rng.min_col).value,
            'max_col': rng.max_col, 'max_row': rng.max_row,
        }
    return merged


@router.post("/import-template")
async def import_template(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """导入模板 Excel：探头-功能配置模板.xlsx"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持 Excel 文件")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    merged_info = _parse_merged_cells(ws)

    # Parse feature groups from row 1 (merged cells)
    # Row 1: merged cells for groups, Row 2: feature names
    group_map = {}  # col_start -> group_name
    group_order = []
    for (row, col), info in sorted(merged_info.items()):
        if row == 1 and info['value']:
            name = str(info['value']).strip()
            if name and '探头-功能配置' not in name:
                group_map[col] = name
                group_order.append((col, name, info['max_col']))

    # Create feature groups
    group_db = {}
    for sort_i, (_, gname, max_col) in enumerate(group_order):
        r = await db.execute(select(FeatureGroup).where(FeatureGroup.name == gname))
        g = r.scalar_one_or_none()
        if not g:
            g = FeatureGroup(name=gname, sort_order=sort_i)
            db.add(g); await db.flush()
        group_db[gname] = g

    # Parse features from row 2 (column by column)
    feature_map = {}  # col -> feature_id
    for col in range(5, ws.max_column + 1):  # E onwards
        val = ws.cell(row=2, column=col).value
        if not val: continue
        name = str(val).strip().replace('\n', ' ')
        # Find which group this column belongs to
        parent_group = None
        for col_start, gname, max_col in group_order:
            if col_start <= col <= max_col:
                parent_group = group_db[gname]
                break
        if not parent_group:
            continue
        r = await db.execute(select(Feature).where(Feature.name == name, Feature.group_id == parent_group.id))
        f = r.scalar_one_or_none()
        if not f:
            f = Feature(name=name, group_id=parent_group.id, sort_order=col)
            db.add(f); await db.flush()
        feature_map[col] = f

    # Parse probe categories and models from rows 3+
    application_map = {}  # name -> Application
    cat_created = {}  # name -> category
    model_created = {}  # (cat_id, model_number) -> probe_model

    for row_idx in range(3, ws.max_row + 1):
        cat_name = ws.cell(row=row_idx, column=1).value  # A: category
        probes_str = ws.cell(row=row_idx, column=2).value  # B: probe models
        apps_str = ws.cell(row=row_idx, column=3).value  # C: applications
        poc_apps_str = ws.cell(row=row_idx, column=4).value  # D: POC applications

        if not cat_name or not probes_str:
            continue
        cat_name = str(cat_name).strip()

        # Create category if needed
        if cat_name not in cat_created:
            r = await db.execute(select(ProbeCategory).where(ProbeCategory.name == cat_name))
            cat = r.scalar_one_or_none()
            if not cat:
                cat = ProbeCategory(name=cat_name, sort_order=row_idx)
                db.add(cat); await db.flush()
            cat_created[cat_name] = cat
        else:
            cat = cat_created[cat_name]

        # Parse probe models (comma/space separated)
        probe_names = [p.strip() for p in probes_str.replace('，', ',').split(',') if p.strip()]
        # Also handle Chinese punctuation
        probe_names = [p for name in probe_names for p in name.split() if p.strip()]

        # Actually, probe models are comma-separated like "SR1-10C, S1-8CX, S1-8CM"
        # Some have spaces mixed in
        probe_entries = []
        for token in probes_str.replace('，', ',').split(','):
            token = token.strip()
            if not token: continue
            # Some entries have multiple models separated by spaces
            for sub in token.split():
                sub = sub.strip(',').strip()
                if sub:
                    probe_entries.append(sub)

        probe_db_list = []
        for pn in probe_entries:
            if not pn or len(pn) < 2: continue
            key = (cat.id, pn)
            if key in model_created:
                probe_db_list.append(model_created[key])
                continue
            r = await db.execute(select(ProbeModel).where(
                ProbeModel.category_id == cat.id, ProbeModel.model_number == pn))
            pm = r.scalar_one_or_none()
            if not pm:
                pm = ProbeModel(category_id=cat.id, model_number=pn, sort_order=len(probe_db_list))
                db.add(pm); await db.flush()
            model_created[key] = pm
            probe_db_list.append(pm)

        # Parse applications — 分别处理常规应用和 POC 应用
        async def _link_apps(apps_str, probe_type):
            if not apps_str: return
            for token in str(apps_str).split(','):
                token = token.strip()
                if not token: continue
                if token in application_map:
                    app = application_map[token]
                else:
                    r = await db.execute(select(Application).where(Application.name == token))
                    app = r.scalar_one_or_none()
                    if not app:
                        app = Application(name=token, sort_order=len(application_map))
                        db.add(app); await db.flush()
                    application_map[token] = app
                # Link category to application with probe_type
                r = await db.execute(select(CategoryApplication).where(
                    CategoryApplication.category_id == cat.id,
                    CategoryApplication.application_id == app.id,
                    CategoryApplication.probe_type == probe_type))
                if not r.scalar_one_or_none():
                    db.add(CategoryApplication(category_id=cat.id, application_id=app.id, probe_type=probe_type))

        await _link_apps(apps_str, "regular")
        await _link_apps(poc_apps_str, "poc")

        # Parse feature support values and create template_features
        for col, feature in feature_map.items():
            val = ws.cell(row=row_idx, column=col).value
            cell_str = str(val).strip() if val else ''
            support = "unsupported"
            excludes = None

            if '√' in cell_str:
                support = "supported"
                # Check for conditions like "√(除早\中\晚孕, 胎心)" or "√(除腹部血管)"
                # Template stores as "supported" with excludes; product level uses "conditional"
                if '(' in cell_str or '（' in cell_str:
                    support = "supported"  # template level: still supported
                    import re
                    cond = re.search(r'[（(](除.*?)[）)]', cell_str)
                    if cond:
                        excl_text = cond.group(1).replace('除', '')
                        exclude_list = [x.strip() for x in excl_text.replace('、', ',').replace('，', ',').split(',') if x.strip()]
                        excludes = json.dumps(exclude_list, ensure_ascii=False)

            if support == "unsupported" and not cell_str:
                continue  # Skip empty cells (no explicit template entry)

            r = await db.execute(select(TemplateFeature).where(
                TemplateFeature.category_id == cat.id,
                TemplateFeature.feature_id == feature.id))
            tf = r.scalar_one_or_none()
            if not tf:
                tf = TemplateFeature(
                    category_id=cat.id, feature_id=feature.id,
                    default_support=support, default_excludes=excludes)
                db.add(tf)
            else:
                tf.default_support = support
                tf.default_excludes = excludes

    await db.commit()
    return {"message": f"导入完成：{len(group_db)} 功能组，{len(feature_map)} 功能，{len(cat_created)} 探头类别，{len(model_created)} 探头型号，{len(application_map)} 应用"}


@router.post("/import-product")
async def import_product(
    file: UploadFile = File(...),
    product_model_id: int = Form(None),
    product_model_name: str = Form(None),
    config_group: str = Form(None),
    db: AsyncSession = Depends(get_db)
):
    """导入产品 Excel：VINNO 6_探头-功能配置.xlsx（支持按 config_group 批量导入）"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持 Excel 文件")

    # Resolve target models
    target_models = []
    if config_group:
        r = await db.execute(select(ProductModel).where(ProductModel.config_group == config_group))
        target_models = r.scalars().all()
        if not target_models:
            raise HTTPException(status_code=400, detail=f"未找到 config_group='{config_group}' 的机型")
    else:
        product_model = None
        if product_model_id:
            r = await db.execute(select(ProductModel).where(ProductModel.id == product_model_id))
            product_model = r.scalar_one_or_none()
        if not product_model and product_model_name:
            r = await db.execute(select(ProductModel).where(ProductModel.name == product_model_name))
            product_model = r.scalar_one_or_none()
        if not product_model:
            raise HTTPException(status_code=400, detail="请指定 product_model_id / product_model_name / config_group")
        target_models = [product_model]

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Load existing features, probe models, categories from DB
    feat_result = await db.execute(select(Feature))
    all_features = {f.name: f for f in feat_result.scalars().all()}

    probe_result = await db.execute(select(ProbeModel))
    all_probes = {(pm.model_number, pm.category_id): pm for pm in probe_result.scalars().all()}

    cat_result = await db.execute(select(ProbeCategory))
    all_cats = {c.name: c for c in cat_result.scalars().all()}

    # Parse feature columns from row 2
    feature_col_map = {}  # col -> Feature
    for col in range(5, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val:
            name = str(val).strip().replace('\n', ' ')
            if name in all_features:
                feature_col_map[col] = all_features[name]

    # Parse probe config rows (starting from row 4) — process per target model
    total_configs = 0
    total_links = 0

    for product_model in target_models:
        created_configs = 0
        created_links = set()

        for row_idx in range(4, ws.max_row + 1):
            priority = str(ws.cell(row=row_idx, column=1).value or '').strip()
            cat_code = str(ws.cell(row=row_idx, column=2).value or '').strip()
            probe_num = str(ws.cell(row=row_idx, column=3).value or '').strip()

            if not probe_num or not cat_code:
                continue

            # Extract category name (remove prefix like "A-", "B-", etc.)
            cat_name = cat_code
            if '-' in cat_code:
                cat_name = cat_code.split('-', 1)[1] if len(cat_code.split('-', 1)) > 1 else cat_code

            # Find category and probe model
            cat = all_cats.get(cat_name)
            if not cat:
                for cn, c in all_cats.items():
                    if cn in cat_name or cat_name in cn:
                        cat = c; break
                if not cat: continue

            # Clean probe number
            probe_num = probe_num.replace('\u00a0', ' ').strip()
            pm = all_probes.get((probe_num, cat.id))
            if not pm:
                for (mn, cid), p in all_probes.items():
                    if mn == probe_num:
                        pm = p; break
                if not pm: continue

            # Create product-probe link
            link_key = (product_model.id, pm.id)
            if link_key not in created_links:
                r = await db.execute(select(ProductProbeModel).where(
                    ProductProbeModel.product_model_id == product_model.id,
                    ProductProbeModel.probe_model_id == pm.id))
                if not r.scalar_one_or_none():
                    ppm = ProductProbeModel(
                        product_model_id=product_model.id,
                        probe_model_id=pm.id,
                        priority=priority if priority else None)
                    db.add(ppm)
                created_links.add(link_key)

            # Parse feature values
            for col, feature in feature_col_map.items():
                val = ws.cell(row=row_idx, column=col).value
                cell_str = str(val).strip() if val else ''
                support = "unsupported"
                excludes = None

                if '√' in cell_str or '✓' in cell_str:
                    support = "supported"
                    if '不支持' in cell_str:
                        support = "unsupported"
                    elif '(' in cell_str or '（' in cell_str:
                        support = "conditional"
                        import re
                        cond = re.search(r'[（(](除.*?)[）)]', cell_str)
                        if cond:
                            excl_text = cond.group(1).replace('除', '')
                            exclude_list = [x.strip() for x in excl_text.replace('、', ',').replace('，', ',').split(',') if x.strip()]
                            excludes = json.dumps(exclude_list, ensure_ascii=False)

                if support == "unsupported" and not cell_str:
                    continue

                # Check template for comparison
                is_overridden = False
                tpl_r = await db.execute(select(TemplateFeature).where(
                    TemplateFeature.category_id == cat.id,
                    TemplateFeature.feature_id == feature.id))
                tpl = tpl_r.scalar_one_or_none()
                if tpl and tpl.default_support != support:
                    is_overridden = True

                # Upsert config
                cfg_r = await db.execute(select(ProductProbeConfig).where(
                    ProductProbeConfig.product_model_id == product_model.id,
                    ProductProbeConfig.probe_model_id == pm.id,
                    ProductProbeConfig.feature_id == feature.id))
                cfg = cfg_r.scalar_one_or_none()
                if not cfg:
                    cfg = ProductProbeConfig(
                        product_model_id=product_model.id,
                        probe_model_id=pm.id,
                        feature_id=feature.id)
                    db.add(cfg)
                cfg.defined_status = support
                cfg.current_status = support
                cfg.defined_excludes = excludes
                cfg.current_excludes = excludes
                cfg.priority = priority if priority else None
                cfg.is_overridden = is_overridden
                created_configs += 1

        total_configs += created_configs
        total_links += len(created_links)

    await db.commit()
    return {"message": f"导入完成：{len(target_models)} 个机型，{total_links} 个探头关联，{total_configs} 条功能配置"}


def _expand_combined_name(name: str) -> list[str]:
    """展开合并的应用名称，如 '早\\中\\晚孕' → ['早孕', '中孕', '晚孕']"""
    name = name.strip()
    # Combined pattern: 早\中\晚孕, 早\中晚孕, etc.
    import re
    # Match patterns like 早\中\晚孕 or 早中晚孕 (combined short names)
    # Common combined prefixes
    parts = re.split(r'[\\/]', name)
    if len(parts) > 1:
        # e.g. "早\中\晚孕" → ["早", "中", "晚孕"]
        # The last part contains the shared suffix
        suffix = ''
        full_names = []
        for i, p in enumerate(parts):
            p = p.strip()
            if not p:
                continue
            if i == len(parts) - 1:
                # Last part has its own suffix
                full_names.append(p)
            else:
                # Suffix is everything after the first char(s) in the last part
                # For "早\中\晚孕": parts = ["早", "中", "晚孕"]
                # "晚孕" → suffix candidates
                last = parts[-1]
                # Extract suffix: "孕" from "晚孕", or just use last as-is
                # Heuristic: if last has more than 1 char, suffix = last[1:]
                if len(last) > 1:
                    suffix = last[1:]
                full_names.append(p + suffix)
        return [n for n in full_names if n]
    return [name]


@router.post("/fill-template-excludes")
async def fill_template_excludes(db: AsyncSession = Depends(get_db)):
    """智能填充模板配置中的应用限制（排除列表）

    算法：
    1. 收集所有已经有排除列表的 TemplateFeature，按 feature_id 聚合
    2. 展开合并名称（如 "早\\中\\晚孕" → ["早孕", "中孕", "晚孕"]）
    3. 对每个排除列表为空的 TemplateFeature：
       a. 获取该类别关联的所有应用名称
       b. 对同一 feature 下其他类别中出现过的排除项，如果该类别也有同名应用，则加入排除
    """
    # 1. Get all template features grouped by feature
    r = await db.execute(select(TemplateFeature))
    all_tfs = r.scalars().all()

    # 2. Build feature -> set of all excluded app names (from features that DO have excludes)
    feat_excludes_map = {}  # feature_id -> set of app names
    for tf in all_tfs:
        if tf.default_excludes and tf.default_support == "supported":
            try:
                ex_list = json.loads(tf.default_excludes)
            except (json.JSONDecodeError, TypeError):
                continue
            if tf.feature_id not in feat_excludes_map:
                feat_excludes_map[tf.feature_id] = set()
            for ex_name in ex_list:
                expanded = _expand_combined_name(ex_name)
                for en in expanded:
                    feat_excludes_map[tf.feature_id].add(en)

    # 3. Get all category -> app names mapping
    r = await db.execute(
        select(CategoryApplication, Application)
        .join(Application, CategoryApplication.application_id == Application.id)
    )
    cat_apps_map = {}  # category_id -> set of app names
    for ca, app in r.all():
        if ca.category_id not in cat_apps_map:
            cat_apps_map[ca.category_id] = set()
        cat_apps_map[ca.category_id].add(app.name)

    # 4. Fill empty excludes
    updated = 0
    for tf in all_tfs:
        if tf.default_support != "supported" or tf.default_excludes is not None:
            continue
        if tf.feature_id not in feat_excludes_map:
            continue
        cat_apps = cat_apps_map.get(tf.category_id, set())
        if not cat_apps:
            continue

        # Find which of the feature's commonly-excluded apps exist in this category
        union_excludes = feat_excludes_map[tf.feature_id]
        applicable = [app for app in union_excludes if app in cat_apps]
        if applicable:
            tf.default_excludes = json.dumps(applicable, ensure_ascii=False)
            updated += 1

    await db.commit()
    return {"message": f"智能填充完成，更新了 {updated} 条模板配置的排除应用列表"}
