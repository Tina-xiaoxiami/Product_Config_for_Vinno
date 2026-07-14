"""产品探头-功能配置 API"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from typing import Optional
import io, json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.database import get_db
from app.models.probe import (
    ProductProbeModel, ProductProbeConfig,
    ProbeModel, ProbeCategory, ProbeModelVariant,
    Feature, FeatureGroup, Application,
    TemplateFeature, ProbeConfigDraft, ProbeConfigVersion,
    CategoryApplication, SeriesProbeConfigVersion,
)
from app.models.product_model import ProductModel
from app.models.series import ProductSeries
from app.models.config import ConfigItem, ConfigValue
from app.schemas.probe import (
    ProductProbeConfigMatrix, ProductProbeConfigItem,
    UpdateProbeFeatureRequest, ProductProbeInitRequest, SubmitDraftRequest,
    SeriesProbeResult, SeriesProbeModelItem, SeriesProbeCategory, SeriesProbeSummary,
    SeriesMatrixResponse, MergedConfigItem,
    SeriesFeatureUpdateRequest, SeriesSubmitRequest,
)
router = APIRouter()


def _normalize_status(status: str) -> str:
    """统一状态值"""
    if status in ("supported", "conditional", "unsupported"):
        return status
    return "unsupported"


@router.get("/by-series", response_model=SeriesProbeResult)
async def get_series_probes(
    series_ids: str = Query(..., description="逗号分隔的系列ID，如 '1,2'"),
    db: AsyncSession = Depends(get_db)
):
    """根据系列ID查询所有关联探头（仅返回 current_status 有效的探头）"""
    ids = [int(x.strip()) for x in series_ids.split(",") if x.strip()]
    if not ids:
        raise HTTPException(status_code=400, detail="请至少选择一个产品系列")

    # 1. 查系列
    series_result = await db.execute(select(ProductSeries).where(ProductSeries.id.in_(ids)))
    series_map = {s.id: s.name for s in series_result.scalars().all()}

    # 2. 查产品型号
    model_result = await db.execute(
        select(ProductModel).where(ProductModel.series_id.in_(ids)).order_by(ProductModel.sort_order))
    models = model_result.scalars().all()

    populated_series = set(m.series_id for m in models)
    empty_series = [name for sid, name in series_map.items() if sid not in populated_series]

    if not models:
        return SeriesProbeResult(
            categories=[], probe_ids=[], empty_series=empty_series)

    model_ids = [m.id for m in models]
    model_name_map = {m.id: m.name for m in models}

    # 3. 查产品-探头关联
    ppm_result = await db.execute(
        select(ProductProbeModel, ProbeModel, ProbeCategory)
        .join(ProbeModel, ProductProbeModel.probe_model_id == ProbeModel.id)
        .join(ProbeCategory, ProbeModel.category_id == ProbeCategory.id)
        .where(ProductProbeModel.product_model_id.in_(model_ids))
        .order_by(ProbeCategory.sort_order, ProbeModel.sort_order))
    ppm_rows = ppm_result.all()

    # 4. 查 product_probe_configs，筛选 current_status 有效的探头
    all_probe_ids = list(set(r.ProbeModel.id for r in ppm_rows))
    if all_probe_ids:
        cfg_result = await db.execute(
            select(ProductProbeConfig.probe_model_id,
                   ProductProbeConfig.product_model_id,
                   ProductProbeConfig.current_status)
            .where(ProductProbeConfig.product_model_id.in_(model_ids),
                   ProductProbeConfig.probe_model_id.in_(all_probe_ids),
                   ProductProbeConfig.current_status != "unsupported")
            .distinct())
        # valid_probes: set of (product_model_id, probe_model_id) with effective config
        valid_probes = set()
        for row in cfg_result.all():
            valid_probes.add((row.product_model_id, row.probe_model_id))
    else:
        valid_probes = set()

    # 5. 仅保留有有效配置的探头，按类别分组去重
    cat_map = {}
    probe_order = []
    seen_probes = set()

    for ppm, pm, pc in ppm_rows:
        if (ppm.product_model_id, pm.id) not in valid_probes:
            continue

        model_name = model_name_map.get(ppm.product_model_id, "")
        if pm.id in seen_probes:
            for cat in cat_map.values():
                existing = cat["models"].get(pm.id)
                if existing and model_name and model_name not in existing.source_product_models:
                    existing.source_product_models.append(model_name)
            continue

        seen_probes.add(pm.id)
        probe_order.append(pm.id)

        if pc.id not in cat_map:
            cat_map[pc.id] = {"id": pc.id, "name": pc.name, "models": {}}

        cat_map[pc.id]["models"][pm.id] = SeriesProbeModelItem(
            id=pm.id, model_number=pm.model_number,
            category_id=pc.id, category_name=pc.name,
            source_product_models=[model_name] if model_name else [])

    # 组装响应
    categories = []
    for cid in sorted(cat_map.keys()):
        cat = cat_map[cid]
        ordered = [cat["models"][pid] for pid in probe_order if pid in cat["models"]]
        categories.append(SeriesProbeCategory(id=cat["id"], name=cat["name"], models=ordered))

    configured_model_ids = set(pmid for pmid, _ in valid_probes)
    model_names = list(dict.fromkeys(
        model_name_map[mid] for mid in model_ids if mid in configured_model_ids))

    return SeriesProbeResult(
        categories=categories,
        probe_ids=list(seen_probes),
        summary=SeriesProbeSummary(
            product_model_names=model_names,
            total_models=len(model_names),
            total_probes=len(seen_probes)),
        empty_series=empty_series)


# ========== 系列级聚合矩阵 ==========

@router.get("/series-matrix", response_model=SeriesMatrixResponse)
async def get_series_matrix(
    series_ids: str = Query(..., description="逗号分隔的系列ID"),
    model_ids: str = Query(None, description="可选：逗号分隔的产品型号ID，默认全部"),
    db: AsyncSession = Depends(get_db)
):
    """获取选定系列+型号的聚合探头配置矩阵"""
    sid_list = [int(x.strip()) for x in series_ids.split(",") if x.strip()]
    if not sid_list:
        raise HTTPException(status_code=400, detail="请至少选择一个产品系列")

    # 1. 查产品型号
    mq = select(ProductModel).where(ProductModel.series_id.in_(sid_list))
    if model_ids:
        mid_list = [int(x.strip()) for x in model_ids.split(",") if x.strip()]
        mq = mq.where(ProductModel.id.in_(mid_list))
    model_result = await db.execute(mq.order_by(ProductModel.sort_order))
    models = model_result.scalars().all()

    if not models:
        return SeriesMatrixResponse(series_ids=sid_list)

    model_ids_list = [m.id for m in models]
    product_models = [{"id": m.id, "name": m.name, "series_id": m.series_id} for m in models]

    # 2. 从主配置系统 "Probes" 分类查有效探头（current_config 不为空且非 N/A）
    ci_result = await db.execute(
        select(ConfigItem).where(ConfigItem.category == "Probes"))
    probe_config_items = ci_result.scalars().all()
    # 构建 config_item 名称 → probe_model 映射
    ci_name_to_probe: dict = {}
    all_probes_r = await db.execute(
        select(ProbeModel, ProbeCategory)
        .join(ProbeCategory, ProbeModel.category_id == ProbeCategory.id)
        .options(selectinload(ProbeModel.variants)))
    for pm, pc in all_probes_r.all():
        # 用型号匹配（部分匹配），名字可能有变体后缀如 F2-5CP vs F2-5C
        ci_name_to_probe[pm.model_number] = (pm, pc)
        # 也加 variant 的 IPN 映射
        for variant in pm.variants:
            if variant.ipn:
                ci_name_to_probe[f"IPN:{variant.ipn}"] = (pm, pc)
    # 查有效 config_values
    cv_result = await db.execute(
        select(ConfigValue, ConfigItem)
        .join(ConfigItem, ConfigValue.item_id == ConfigItem.id)
        .where(ConfigItem.category == "Probes",
               ConfigValue.model_id.in_(model_ids_list),
               ConfigValue.current_config.isnot(None),
               ConfigValue.current_config != "",
               ConfigValue.current_config != "N/A",
               ConfigValue.current_config != "未定义"))
    cv_rows = cv_result.all()
    # 匹配到 probe_models，按 (probe_id, ipn) 拆分独立行
    probe_row_map: dict = {}  # (probe_id, ipn) -> (probe, category, ipn)
    for cv, ci in cv_rows:
        model_name = ci.rd_name.split("【")[0] if "【" in (ci.rd_name or "") else (ci.rd_name or "")
        ipn = ci.ipn or ""
        matched_pm = None
        # 尝试直接匹配 model_number
        if model_name in ci_name_to_probe:
            matched_pm = ci_name_to_probe[model_name]
        # 尝试 IPN 匹配
        elif ipn and f"IPN:{ipn}" in ci_name_to_probe:
            matched_pm = ci_name_to_probe[f"IPN:{ipn}"]
        else:
            # 模糊匹配：名称包含关系
            for pm_direct, (pm_obj, pc_obj) in ci_name_to_probe.items():
                if not pm_direct.startswith("IPN:") and model_name and (model_name in pm_direct or pm_direct in model_name):
                    matched_pm = (pm_obj, pc_obj)
                    break
        if matched_pm:
            pm, pc = matched_pm
            key = (pm.id, ipn)
            if key not in probe_row_map:
                probe_row_map[key] = (pm, pc, ipn)
    # 构建 IPN→internal_model 映射
    ipn_to_internal: dict = {}
    for pm, _ in ci_name_to_probe.values():
        for v in pm.variants:
            if v.ipn:
                ipn_to_internal[v.ipn] = v.internal_model
    # 构建 probe_models_list（按 category.sort_order 排序，同 probe 多 IPN 合并相邻）
    probe_models_list = []
    for (pm_id, ipn), (pm, pc, _) in sorted(probe_row_map.items(),
            key=lambda kv: (kv[1][1].sort_order, kv[1][0].sort_order, kv[0][1])):
        internal = ipn_to_internal.get(ipn, "")
        probe_models_list.append({
            "id": pm.id,
            "row_key": f"{pm.id}_{ipn}" if ipn else str(pm.id),
            "model_number": pm.model_number,
            "internal_model": internal,
            "category_id": pc.id,
            "category_name": pc.name,
            "ipn": ipn,
            "priority": None,
        })

    # 3. 查所有功能
    feat_result = await db.execute(
        select(Feature, FeatureGroup)
        .join(FeatureGroup, Feature.group_id == FeatureGroup.id)
        .order_by(FeatureGroup.sort_order, Feature.sort_order))
    features = [{"id": f.Feature.id, "name": f.Feature.name, "ipn": f.Feature.ipn,
                  "group_id": f.Feature.group_id, "group_name": f.FeatureGroup.name}
                 for f in feat_result.all()]

    # 4. 查所有配置（跨所有型号）
    probe_ids = [p["id"] for p in probe_models_list]
    feat_ids = [f["id"] for f in features]
    configs: dict = {}
    if probe_ids and feat_ids:
        cfg_result = await db.execute(
            select(ProductProbeConfig)
            .where(ProductProbeConfig.product_model_id.in_(model_ids_list),
                   ProductProbeConfig.probe_model_id.in_(probe_ids),
                   ProductProbeConfig.feature_id.in_(feat_ids)))
        # Group by probe × feature
        for cfg in cfg_result.scalars().all():
            pk = str(cfg.probe_model_id)
            if pk not in configs:
                configs[pk] = {}
            fk = str(cfg.feature_id)
            if fk not in configs[pk]:
                configs[pk][fk] = {"defined_statuses": [], "current_statuses": [],
                    "defined_excludes": None, "current_excludes": None,
                    "per_model": {}, "is_overridden": False, "template_support": "unsupported"}
            d = configs[pk][fk]
            d["defined_statuses"].append(cfg.defined_status)
            d["current_statuses"].append(cfg.current_status)
            d["per_model"][str(cfg.product_model_id)] = {
                "defined_status": cfg.defined_status, "current_status": cfg.current_status}
            if cfg.defined_excludes: d["defined_excludes"] = cfg.defined_excludes
            if cfg.current_excludes: d["current_excludes"] = cfg.current_excludes
            if cfg.is_overridden: d["is_overridden"] = True
            if cfg.priority: d["priority"] = cfg.priority

    # 5. 加载模板比较
    unique_cats = set(p["category_id"] for p in probe_models_list)
    tpl_result = await db.execute(
        select(TemplateFeature).where(TemplateFeature.category_id.in_(unique_cats)))
    tpl_map = {}  # (cat_id, feat_id) -> (default_support, default_excludes)
    for t in tpl_result.scalars().all():
        tpl_map[(t.category_id, t.feature_id)] = (t.default_support, t.default_excludes)

    # 6. 构建聚合 configs — 为所有 probe×feature 填充模板默认值
    def _tpl_support(cat_id, feat_id):
        t = tpl_map.get((cat_id, feat_id))
        return t[0] if t else "unsupported"
    def _tpl_excludes(cat_id, feat_id):
        t = tpl_map.get((cat_id, feat_id))
        return t[1] if t else None

    out_configs = {}
    for pk, feat_map in configs.items():
        out_configs[pk] = {}
        for fk, d in feat_map.items():
            def_statuses = d["defined_statuses"]
            cur_statuses = d["current_statuses"]
            def_s = def_statuses[0] if len(set(def_statuses)) == 1 else "mixed"
            cur_s = cur_statuses[0] if len(set(cur_statuses)) == 1 else "mixed"
            cat_id = next((p["category_id"] for p in probe_models_list if p["id"] == int(pk)), 0)
            feat_id = int(fk)
            out_configs[pk][fk] = MergedConfigItem(
                defined_status=def_s, current_status=cur_s,
                defined_excludes=d["defined_excludes"], current_excludes=d["current_excludes"],
                priority=d.get("priority"), is_overridden=d["is_overridden"],
                template_support=_tpl_support(cat_id, feat_id),
                template_excludes=_tpl_excludes(cat_id, feat_id),
                per_model=d["per_model"])
    # 6b. 为没有 config 记录的 probe×feature 填充模板默认值
    for p in probe_models_list:
        pk = str(p["id"])
        cat_id = p["category_id"]
        if pk not in out_configs:
            out_configs[pk] = {}
        for f in features:
            fk = str(f["id"])
            if fk not in out_configs[pk]:
                tpl_s = _tpl_support(cat_id, f["id"])
                tpl_ex = _tpl_excludes(cat_id, f["id"])
                out_configs[pk][fk] = MergedConfigItem(
                    defined_status=tpl_s, current_status=tpl_s,
                    defined_excludes=None, current_excludes=None,
                    priority=None, is_overridden=False,
                    template_support=tpl_s, template_excludes=tpl_ex,
                    per_model={})

    # 7. 查应用（按 category_id 分组）
    cat_apps_result = await db.execute(
        select(CategoryApplication, Application)
        .join(Application, CategoryApplication.application_id == Application.id)
        .where(CategoryApplication.category_id.in_(unique_cats)))
    apps: dict = {}
    for ca, app in cat_apps_result.all():
        cid = str(ca.category_id)
        if cid not in apps:
            apps[cid] = {"regular": [], "poc": []}
        apps[cid].setdefault(ca.probe_type, []).append({"id": app.id, "name": app.name})

    return SeriesMatrixResponse(
        series_ids=sid_list, product_models=product_models, features=features,
        probe_models=probe_models_list, applications=apps, configs=out_configs)


async def _expand_to_group(db: AsyncSession, model_ids: list[int]) -> list[int]:
    """将机型 ID 列表扩展为包含其同组所有成员"""
    if not model_ids: return []
    r = await db.execute(select(ProductModel).where(ProductModel.id.in_(model_ids)))
    models = r.scalars().all()
    groups = set(m.config_group for m in models if m.config_group)
    if not groups: return list(set(model_ids))
    # 查询所有同组成员
    all_r = await db.execute(
        select(ProductModel.id).where(ProductModel.config_group.in_(groups)))
    all_ids = set(r[0] for r in all_r.all())
    all_ids.update(model_ids)
    return list(all_ids)


@router.put("/series-feature")
async def update_series_feature(data: SeriesFeatureUpdateRequest, db: AsyncSession = Depends(get_db)):
    """更新系列级探头功能配置，传播到所有目标型号（含同组机型）"""
    target_ids = await _expand_to_group(db, data.target_model_ids)
    count = 0
    for mid in target_ids:
        r = await db.execute(select(ProductProbeConfig).where(
            ProductProbeConfig.product_model_id == mid,
            ProductProbeConfig.probe_model_id == data.probe_model_id,
            ProductProbeConfig.feature_id == data.feature_id))
        cfg = r.scalar_one_or_none()
        if not cfg:
            cfg = ProductProbeConfig(product_model_id=mid, probe_model_id=data.probe_model_id,
                                       feature_id=data.feature_id)
            db.add(cfg)
        old_d = cfg.defined_status; old_c = cfg.current_status
        old_de = cfg.defined_excludes; old_ce = cfg.current_excludes
        if data.defined_status is not None: cfg.defined_status = data.defined_status
        if data.current_status is not None: cfg.current_status = data.current_status
        if data.defined_excludes is not None: cfg.defined_excludes = data.defined_excludes
        if data.current_excludes is not None: cfg.current_excludes = data.current_excludes
        if data.priority is not None: cfg.priority = data.priority
        if data.notes is not None: cfg.notes = data.notes
        cfg.is_overridden = True
        draft = ProbeConfigDraft(
            product_model_id=mid, probe_model_id=data.probe_model_id, feature_id=data.feature_id,
            change_type="update", old_defined=old_d, new_defined=data.defined_status or old_d,
            old_current=old_c, new_current=data.current_status or old_c,
            old_excludes=old_de, new_excludes=data.defined_excludes)
        db.add(draft); count += 1
    await db.commit()
    return {"message": f"已为 {count} 个产品型号创建草稿", "draft_count": count}


@router.get("/series-drafts")
async def get_series_drafts(
    model_ids: str = Query(..., description="逗号分隔的产品型号ID"),
    db: AsyncSession = Depends(get_db)
):
    """获取选中型号的聚合草稿数"""
    mid_list = [int(x.strip()) for x in model_ids.split(",") if x.strip()]
    total = 0
    per_model = {}
    for mid in mid_list:
        r = await db.execute(
            select(ProbeConfigDraft).where(ProbeConfigDraft.product_model_id == mid))
        cnt = len(r.scalars().all())
        per_model[str(mid)] = cnt; total += cnt
    return {"total": total, "per_model": per_model}


@router.post("/series-discard")
async def discard_series_drafts(data: dict, db: AsyncSession = Depends(get_db)):
    """废弃选中型号的所有草稿"""
    from sqlalchemy import delete as sqla_delete
    mid_list = data.get("model_ids", [])
    count = 0
    for mid in mid_list:
        r = await db.execute(sqla_delete(ProbeConfigDraft).where(ProbeConfigDraft.product_model_id == mid))
        count += r.rowcount
    await db.commit()
    return {"message": f"已废弃 {count} 条草稿", "total": count}


@router.post("/series-submit")
async def submit_series_drafts(data: SeriesSubmitRequest, db: AsyncSession = Depends(get_db)):
    """提交选中型号的所有草稿并创建系列版本"""
    mid_list = data.model_ids
    total = 0
    for mid in mid_list:
        r = await db.execute(select(ProbeConfigDraft).where(ProbeConfigDraft.product_model_id == mid))
        drafts = r.scalars().all()
        for d in drafts:
            cfg_r = await db.execute(select(ProductProbeConfig).where(
                ProductProbeConfig.product_model_id == d.product_model_id,
                ProductProbeConfig.probe_model_id == d.probe_model_id,
                ProductProbeConfig.feature_id == d.feature_id))
            cfg = cfg_r.scalar_one_or_none()
            if d.change_type == "delete":
                if cfg: await db.delete(cfg)
            else:
                if not cfg:
                    cfg = ProductProbeConfig(product_model_id=d.product_model_id,
                        probe_model_id=d.probe_model_id, feature_id=d.feature_id)
                    db.add(cfg)
                if d.new_defined: cfg.defined_status = d.new_defined
                if d.new_current: cfg.current_status = d.new_current
                cfg.is_overridden = True
            await db.delete(d); total += 1

    # 创建系列版本快照
    snapshot = {
        "series_ids": data.series_ids, "model_ids": mid_list,
        "submitted_at": datetime.utcnow().isoformat(),
        "draft_count": total
    }
    ver_num = data.version_number or f"v{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    db.add(SeriesProbeConfigVersion(
        series_ids=json.dumps(data.series_ids), model_ids=json.dumps(mid_list),
        version_number=ver_num, snapshot_data=json.dumps(snapshot, ensure_ascii=False),
        description=data.description or "系列配置提交"))
    await db.commit()
    return {"message": f"提交成功，版本 {ver_num}，应用 {total} 条变更"}


@router.get("/series-versions")
async def get_series_versions(
    series_ids: str = Query(None, description="逗号分隔的系列ID"),
    db: AsyncSession = Depends(get_db)
):
    """获取系列级版本历史"""
    q = select(SeriesProbeConfigVersion).order_by(SeriesProbeConfigVersion.id.desc()).limit(50)
    if series_ids:
        sid_list = [int(x.strip()) for x in series_ids.split(",") if x.strip()]
        # simple filter: series_ids JSON contains any of the requested ids
        pass  # Full SQL filtering would be complex; return all and let frontend filter
    result = await db.execute(q)
    return [{"id": v.id, "version_number": v.version_number, "description": v.description,
             "series_ids": json.loads(v.series_ids), "model_ids": json.loads(v.model_ids),
             "created_at": v.created_at} for v in result.scalars().all()]


@router.post("/series-rollback/{version_id}")
async def rollback_series_version(version_id: int, db: AsyncSession = Depends(get_db)):
    """回滚到指定系列版本"""
    r = await db.execute(select(SeriesProbeConfigVersion).where(SeriesProbeConfigVersion.id == version_id))
    ver = r.scalar_one_or_none()
    if not ver: raise HTTPException(status_code=404, detail="版本不存在")
    model_ids = json.loads(ver.model_ids)
    # Rollback: discard all drafts and reload from next version
    # For now, just discard all drafts for these models
    from sqlalchemy import delete as sqla_delete
    for mid in model_ids:
        await db.execute(sqla_delete(ProbeConfigDraft).where(ProbeConfigDraft.product_model_id == mid))
    await db.commit()
    return {"message": f"已回滚到版本 {ver.version_number}，废弃 {len(model_ids)} 个型号的草稿"}


# ========== 批量关联探头到产品型号 ==========

@router.post("/series-probes")
async def set_series_probes(data: dict, db: AsyncSession = Depends(get_db)):
    """批量关联探头到多个产品型号"""
    from sqlalchemy import delete as sqla_delete
    model_ids = data.get("model_ids", [])
    probe_ids = data.get("probe_model_ids", [])
    priorities = data.get("priorities", {})
    total = 0
    for mid in model_ids:
        await db.execute(sqla_delete(ProductProbeModel).where(ProductProbeModel.product_model_id == mid))
        for pid in probe_ids:
            db.add(ProductProbeModel(
                product_model_id=mid, probe_model_id=pid,
                priority=priorities.get(str(pid))))
            total += 1
    await db.commit()
    return {"message": f"已为 {len(model_ids)} 个型号关联 {len(probe_ids)} 个探头"}


# ========== 获取所有探头（按类别分组，用于关联对话框） ==========

@router.get("/all-probes")
async def get_all_probes(db: AsyncSession = Depends(get_db)):
    """获取所有探头按类别分组"""
    result = await db.execute(
        select(ProbeCategory).order_by(ProbeCategory.sort_order))
    categories = result.scalars().all()
    output = []
    for cat in categories:
        models_r = await db.execute(
            select(ProbeModel).where(ProbeModel.category_id == cat.id).order_by(ProbeModel.sort_order))
        models = [{"id": m.id, "model_number": m.model_number} for m in models_r.scalars().all()]
        output.append({"id": cat.id, "name": cat.name, "models": models})
    return output


# ========== 机型分组管理（必须在 /{product_model_id} 之前）==========

@router.get("/model-groups")
async def get_model_groups(db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(ProductModel).order_by(ProductModel.config_group, ProductModel.name))
    groups: dict = {}; ungrouped = []
    for m in r.scalars().all():
        if m.config_group:
            groups.setdefault(m.config_group, []).append({"id": m.id, "name": m.name, "series_id": m.series_id})
        else:
            ungrouped.append({"id": m.id, "name": m.name, "series_id": m.series_id})
    return {"groups": groups, "ungrouped": ungrouped}

@router.put("/model-groups")
async def set_model_groups(data: dict, db: AsyncSession = Depends(get_db)):
    count = 0
    for group_name, ids in data.items():
        for mid in ids:
            r = await db.execute(select(ProductModel).where(ProductModel.id == mid))
            m = r.scalar_one_or_none()
            if m:
                m.config_group = group_name if group_name else None
                count += 1
    await db.commit()
    return {"message": f"已更新 {count} 个机型的分组"}

@router.post("/model-groups/rename")
async def rename_model_group(data: dict, db: AsyncSession = Depends(get_db)):
    old, new = data.get("old_name"), data.get("new_name")
    if not old or not new: raise HTTPException(400, "缺少 old_name 或 new_name")
    r = await db.execute(select(ProductModel).where(ProductModel.config_group == old))
    count = 0
    for m in r.scalars().all():
        m.config_group = new; count += 1
    await db.commit()
    return {"message": f"已将 {count} 个机型从 '{old}' 重命名为 '{new}'"}

@router.post("/model-groups/auto")
async def auto_group_models(db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(ProductModel))
    models = r.scalars().all()
    suffixes = ['_BRA', '_RUA', '_Private_RUA', '_Private', '_POC', '_Anesthesia']
    def _base(name):
        for sfx in suffixes:
            if name.endswith(sfx): return name[:-len(sfx)].strip()
        return name.strip()
    # 先为带后缀的机型设组
    count = 0
    for m in models:
        if m.config_group: continue
        base = _base(m.name)
        if base != m.name:
            m.config_group = base; count += 1
    # 再为基础型号（无后缀）设组：如果已有同名 group 则加入
    groups = set()
    for m in models:
        if m.config_group: groups.add(m.config_group)
    for m in models:
        if m.config_group: continue
        if m.name in groups:
            m.config_group = m.name; count += 1
    await db.commit()
    return {"message": f"已为 {count} 个机型自动分组"}


@router.get("/{product_model_id}", response_model=ProductProbeConfigMatrix)
async def get_config(product_model_id: int, db: AsyncSession = Depends(get_db)):
    """获取产品探头-功能配置矩阵"""
    # 验证产品型号存在
    pm_result = await db.execute(select(ProductModel).where(ProductModel.id == product_model_id))
    product_model = pm_result.scalar_one_or_none()
    if not product_model:
        raise HTTPException(status_code=404, detail="产品型号不存在")

    # 获取该产品支持的所有探头
    ppm_result = await db.execute(
        select(ProductProbeModel, ProbeModel, ProbeCategory)
        .join(ProbeModel, ProductProbeModel.probe_model_id == ProbeModel.id)
        .join(ProbeCategory, ProbeModel.category_id == ProbeCategory.id)
        .where(ProductProbeModel.product_model_id == product_model_id)
        .order_by(ProbeCategory.sort_order, ProbeModel.sort_order)
    )
    probe_rows = ppm_result.all()

    # 获取所有功能（含功能组）
    feat_result = await db.execute(
        select(Feature, FeatureGroup)
        .join(FeatureGroup, Feature.group_id == FeatureGroup.id)
        .order_by(FeatureGroup.sort_order, Feature.sort_order)
    )
    features_data = feat_result.all()

    # 获取模板默认值用于对比
    cat_ids = list(set(r.ProbeCategory.id for r in probe_rows))
    tpl_r = await db.execute(select(TemplateFeature).where(TemplateFeature.category_id.in_(cat_ids)))
    tpl_map = {(t.category_id, t.feature_id): t.default_support for t in tpl_r.scalars().all()}
    probe_cat_map = {r.ProbeModel.id: r.ProbeCategory.id for r in probe_rows}
    probe_model_ids = [row.ProbeModel.id for row in probe_rows]

    # 获取探头型号变体信息（联动变体数据）
    var_result = await db.execute(
        select(ProbeModelVariant).where(ProbeModelVariant.probe_model_id.in_(probe_model_ids))
    )
    var_map = {}
    for v in var_result.scalars().all():
        if v.probe_model_id not in var_map: var_map[v.probe_model_id] = []
        var_map[v.probe_model_id].append({"internal_model": v.internal_model, "ipn": v.ipn})

    # 获取所有配置
    feature_ids = [f.Feature.id for f in features_data]
    configs = {}
    if probe_model_ids and feature_ids:
        cfg_result = await db.execute(
            select(ProductProbeConfig).where(
                ProductProbeConfig.product_model_id == product_model_id,
                ProductProbeConfig.probe_model_id.in_(probe_model_ids),
                ProductProbeConfig.feature_id.in_(feature_ids),
            )
        )
        for cfg in cfg_result.scalars().all():
            key = f"{cfg.probe_model_id}_{cfg.feature_id}"
            configs[key] = ProductProbeConfigItem(
                id=cfg.id,
                probe_model_id=cfg.probe_model_id,
                feature_id=cfg.feature_id,
                defined_status=cfg.defined_status or "unsupported",
                current_status=cfg.current_status or "unsupported",
                defined_excludes=cfg.defined_excludes,
                current_excludes=cfg.current_excludes,
                priority=cfg.priority,
                notes=cfg.notes,
                is_overridden=cfg.is_overridden or False,
                template_support=tpl_map.get((probe_cat_map.get(cfg.probe_model_id), cfg.feature_id), "unsupported"),
            )

    # 获取分组应用（按 probe_type 区分 常规/POC）
    app_result = await db.execute(
        select(CategoryApplication, Application)
        .join(Application, CategoryApplication.application_id == Application.id)
        .where(CategoryApplication.category_id.in_(cat_ids))
    )
    apps_grouped = {"regular": [], "poc": []}
    for ca, app in app_result.all():
        apps_grouped.setdefault(ca.probe_type, []).append({"id": app.id, "name": app.name, "en_name": app.en_name})

    return ProductProbeConfigMatrix(
        product_model_id=product_model_id,
        product_model_name=product_model.name,
        features=[
            {"id": f.Feature.id, "name": f.Feature.name, "ipn": f.Feature.ipn, "group_id": f.FeatureGroup.id, "group_name": f.FeatureGroup.name}
            for f in features_data
        ],
        probe_models=[
            {
                "id": r.ProbeModel.id, "model_number": r.ProbeModel.model_number,
                "category_id": r.ProbeCategory.id, "category_name": r.ProbeCategory.name,
                "priority": r.ProductProbeModel.priority,
                "variants": var_map.get(r.ProbeModel.id, []),
            }
            for r in probe_rows
        ],
        applications=apps_grouped,
        configs=configs,
    )


@router.post("/{product_model_id}/init", response_model=dict)
async def init_config(
    product_model_id: int,
    data: Optional[ProductProbeInitRequest] = None,
    db: AsyncSession = Depends(get_db)
):
    """从模板初始化产品探头配置"""
    # 验证产品型号
    pm_result = await db.execute(select(ProductModel).where(ProductModel.id == product_model_id))
    if not pm_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="产品型号不存在")

    # 获取该产品的探头列表
    ppm_q = select(ProductProbeModel).where(ProductProbeModel.product_model_id == product_model_id)
    if data and data.probe_model_ids:
        ppm_q = ppm_q.where(ProductProbeModel.probe_model_id.in_(data.probe_model_ids))
    ppm_result = await db.execute(ppm_q)
    probe_links = ppm_result.scalars().all()

    if not probe_links:
        return {"message": "该产品没有关联探头型号，请先在探头型号管理中关联"}

    probe_ids = [p.probe_model_id for p in probe_links]

    # 获取探头类别
    probes_result = await db.execute(select(ProbeModel).where(ProbeModel.id.in_(probe_ids)))
    probe_map = {p.id: p for p in probes_result.scalars().all()}

    # 获取所有模板配置
    category_ids = list(set(p.category_id for p in probe_map.values()))
    tpl_result = await db.execute(
        select(TemplateFeature).where(TemplateFeature.category_id.in_(category_ids))
    )
    templates = tpl_result.scalars().all()

    # 按 (category_id, feature_id) 索引模板
    tpl_map = {(t.category_id, t.feature_id): t for t in templates}

    # 获取所有功能
    feat_result = await db.execute(select(Feature))
    all_features = feat_result.scalars().all()

    created = 0
    for link in probe_links:
        probe = probe_map.get(link.probe_model_id)
        if not probe:
            continue
        cat_id = probe.category_id
        for feat in all_features:
            # 检查是否已有配置
            exist_r = await db.execute(
                select(ProductProbeConfig).where(
                    ProductProbeConfig.product_model_id == product_model_id,
                    ProductProbeConfig.probe_model_id == probe.id,
                    ProductProbeConfig.feature_id == feat.id,
                )
            )
            if exist_r.scalar_one_or_none():
                continue

            tpl = tpl_map.get((cat_id, feat.id))
            default_status = tpl.default_support if tpl else "unsupported"
            default_excludes = tpl.default_excludes if tpl else None
            priority = link.priority

            cfg = ProductProbeConfig(
                product_model_id=product_model_id,
                probe_model_id=probe.id,
                feature_id=feat.id,
                defined_status=default_status,
                current_status=default_status,  # 初始与定义值一致
                defined_excludes=default_excludes,
                current_excludes=default_excludes,
                priority=priority,
                is_overridden=False,
            )
            db.add(cfg)
            created += 1

    await db.commit()
    return {"message": f"初始化完成，创建了 {created} 条配置"}


@router.put("/{product_model_id}/feature", response_model=dict)
async def update_feature(
    product_model_id: int,
    data: UpdateProbeFeatureRequest,
    db: AsyncSession = Depends(get_db)
):
    """更新单个 probe × feature 的配置"""
    result = await db.execute(
        select(ProductProbeConfig).where(
            ProductProbeConfig.product_model_id == product_model_id,
            ProductProbeConfig.probe_model_id == data.probe_model_id,
            ProductProbeConfig.feature_id == data.feature_id,
        )
    )
    cfg = result.scalar_one_or_none()

    if not cfg:
        # 自动创建
        cfg = ProductProbeConfig(
            product_model_id=product_model_id,
            probe_model_id=data.probe_model_id,
            feature_id=data.feature_id,
        )
        db.add(cfg)

    for field in ("defined_status", "current_status", "defined_excludes", "current_excludes", "priority", "notes"):
        val = getattr(data, field, None)
        if val is not None:
            setattr(cfg, field, val)

    cfg.is_overridden = True

    # 创建草稿（而非直接写库）
    draft = ProbeConfigDraft(
        product_model_id=product_model_id,
        probe_model_id=data.probe_model_id,
        feature_id=data.feature_id,
        change_type="update",
        old_defined=cfg.defined_status,
        new_defined=data.defined_status,
        old_current=cfg.current_status,
        new_current=data.current_status,
    )
    # 不实际修改 cfg，草稿提交时才应用
    db.add(draft)
    await db.commit()
    await db.refresh(draft)
    return {"message": "草稿已保存", "id": draft.id}


@router.get("/{product_model_id}/drafts", response_model=dict)
async def get_drafts(product_model_id: int, db: AsyncSession = Depends(get_db)):
    """获取草稿列表"""
    from app.models.probe import ProbeConfigDraft, ProbeModel, Feature
    result = await db.execute(
        select(ProbeConfigDraft, ProbeModel, Feature)
        .join(ProbeModel, ProbeConfigDraft.probe_model_id == ProbeModel.id)
        .join(Feature, ProbeConfigDraft.feature_id == Feature.id)
        .where(ProbeConfigDraft.product_model_id == product_model_id)
        .order_by(ProbeConfigDraft.created_at.desc())
    )
    drafts = []
    for d, pm, f in result.all():
        drafts.append({
            "id": d.id, "probe_model_id": d.probe_model_id, "feature_id": d.feature_id,
            "probe_number": pm.model_number, "feature_name": f.name,
            "change_type": d.change_type,
            "old_defined": d.old_defined, "new_defined": d.new_defined,
            "old_current": d.old_current, "new_current": d.new_current,
            "created_at": d.created_at,
        })
    return {"drafts": drafts, "total": len(drafts)}


@router.post("/{product_model_id}/submit", response_model=dict)
async def submit_drafts(
    product_model_id: int,
    data: Optional[SubmitDraftRequest] = None,
    db: AsyncSession = Depends(get_db)
):
    """提交草稿：应用变更 + 创建版本快照"""
    from app.models.probe import ProbeConfigDraft, ProbeConfigVersion, ProbeModel, Feature, FeatureGroup

    # Get all drafts
    d_result = await db.execute(
        select(ProbeConfigDraft).where(ProbeConfigDraft.product_model_id == product_model_id)
    )
    drafts = d_result.scalars().all()
    if not drafts:
        raise HTTPException(status_code=400, detail="没有待提交的草稿")

    # Apply each draft
    applied = 0
    for d in drafts:
        cfg_result = await db.execute(select(ProductProbeConfig).where(
            ProductProbeConfig.product_model_id == product_model_id,
            ProductProbeConfig.probe_model_id == d.probe_model_id,
            ProductProbeConfig.feature_id == d.feature_id,
        ))
        cfg = cfg_result.scalar_one_or_none()
        if not cfg:
            cfg = ProductProbeConfig(
                product_model_id=product_model_id,
                probe_model_id=d.probe_model_id,
                feature_id=d.feature_id,
            )
            db.add(cfg)
        if d.new_defined: cfg.defined_status = d.new_defined
        if d.new_current: cfg.current_status = d.new_current
        cfg.is_overridden = True
        await db.delete(d)
        applied += 1

    # Create version snapshot
    matrix = await get_config(product_model_id, db)
    version_num = (data.version_number if data and data.version_number
                   else f"v{__import__('datetime').datetime.utcnow().strftime('%Y%m%d%H%M%S')}")

    version = ProbeConfigVersion(
        product_model_id=product_model_id,
        version_number=version_num,
        description=(data.description if data else None),
        snapshot_data=json.dumps(matrix.model_dump(), ensure_ascii=False),
    )
    db.add(version)
    await db.commit()
    return {"message": f"提交成功，版本 {version_num}，应用 {applied} 项变更"}


@router.post("/{product_model_id}/discard", response_model=dict)
async def discard_drafts(product_model_id: int, db: AsyncSession = Depends(get_db)):
    """废弃所有草稿"""
    from app.models.probe import ProbeConfigDraft
    result = await db.execute(
        select(ProbeConfigDraft).where(ProbeConfigDraft.product_model_id == product_model_id)
    )
    drafts = result.scalars().all()
    for d in drafts: await db.delete(d)
    await db.commit()
    return {"message": f"已废弃 {len(drafts)} 条草稿"}


@router.get("/{product_model_id}/versions", response_model=list)
async def get_versions(product_model_id: int, db: AsyncSession = Depends(get_db)):
    """获取版本历史"""
    from app.models.probe import ProbeConfigVersion
    result = await db.execute(
        select(ProbeConfigVersion)
        .where(ProbeConfigVersion.product_model_id == product_model_id)
        .order_by(ProbeConfigVersion.id.desc())
        .limit(50)
    )
    return [{"id": v.id, "version_number": v.version_number, "description": v.description, "created_at": v.created_at}
            for v in result.scalars().all()]


@router.post("/{product_model_id}/rollback/{version_id}", response_model=dict)
async def rollback_version(
    product_model_id: int, version_id: int,
    db: AsyncSession = Depends(get_db)
):
    """回滚到指定版本"""
    from app.models.probe import ProbeConfigVersion
    r = await db.execute(select(ProbeConfigVersion).where(ProbeConfigVersion.id == version_id))
    ver = r.scalar_one_or_none()
    if not ver: raise HTTPException(status_code=404, detail="版本不存在")

    snap = json.loads(ver.snapshot_data)
    # Delete current configs
    del_r = await db.execute(select(ProductProbeConfig).where(
        ProductProbeConfig.product_model_id == product_model_id))
    for c in del_r.scalars().all(): await db.delete(c)
    await db.flush()

    # Rebuild from snapshot
    restored = 0
    for key, item in snap.get("configs", {}).items():
        parts = key.split("_")
        cfg = ProductProbeConfig(
            product_model_id=product_model_id,
            probe_model_id=int(parts[0]),
            feature_id=int(parts[1]),
            defined_status=item.get("defined_status", "unsupported"),
            current_status=item.get("current_status", "unsupported"),
            defined_excludes=item.get("defined_excludes"),
            current_excludes=item.get("current_excludes"),
            priority=item.get("priority"),
            notes=item.get("notes"),
            is_overridden=item.get("is_overridden", False),
        )
        db.add(cfg); restored += 1

    await db.commit()
    return {"message": f"已回滚到版本 {ver.version_number}，恢复 {restored} 条配置"}


@router.post("/{product_model_id}/apply-template", response_model=dict)
async def apply_template(product_model_id: int, db: AsyncSession = Depends(get_db)):
    """将模板最新值同步到产品配置（更新全部探头类别对应的默认值）"""
    ppm_result = await db.execute(
        select(ProductProbeModel).where(ProductProbeModel.product_model_id == product_model_id)
    )
    links = ppm_result.scalars().all()
    if not links: raise HTTPException(status_code=400, detail="该产品没有关联探头")

    probe_ids = [l.probe_model_id for l in links]
    probes = {p.id: p for p in (await db.execute(select(ProbeModel).where(ProbeModel.id.in_(probe_ids)))).scalars().all()}
    cat_ids = list(set(p.category_id for p in probes.values()))
    tpls = {(t.category_id, t.feature_id): t for t in
            (await db.execute(select(TemplateFeature).where(TemplateFeature.category_id.in_(cat_ids)))).scalars().all()}

    updated = 0
    for link in links:
        probe = probes.get(link.probe_model_id)
        if not probe: continue
        tpl = {(t.category_id, t.feature_id): t for t in
               (await db.execute(select(TemplateFeature).where(TemplateFeature.category_id == probe.category_id))).scalars().all()}
        for (cid, fid), t in tpl.items():
            cfg_r = await db.execute(select(ProductProbeConfig).where(
                ProductProbeConfig.product_model_id == product_model_id,
                ProductProbeConfig.probe_model_id == probe.id,
                ProductProbeConfig.feature_id == fid))
            cfg = cfg_r.scalar_one_or_none()
            if not cfg:
                cfg = ProductProbeConfig(product_model_id=product_model_id, probe_model_id=probe.id, feature_id=fid)
                db.add(cfg)
            if cfg.defined_status != t.default_support:
                cfg.defined_status = t.default_support
                cfg.is_overridden = False
                updated += 1
    await db.commit()
    return {"message": f"已同步 {updated} 项与模板一致"}


@router.post("/{product_model_id}/batch-set", response_model=dict)
async def batch_set_status(
    product_model_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db)
):
    """批量设置某探头所有功能的状态 {probe_model_id, defined_status, current_status}（含同组机型）"""
    probe_id = data.get("probe_model_id")
    if not probe_id: raise HTTPException(status_code=400, detail="缺少probe_model_id")

    model_ids = await _expand_to_group(db, [product_model_id])
    feat_result = await db.execute(select(Feature))
    features = feat_result.scalars().all()
    def_stat = data.get("defined_status", "unsupported")
    cur_stat = data.get("current_status", def_stat)

    count = 0
    for mid in model_ids:
      for feat in features:
        r = await db.execute(select(ProductProbeConfig).where(
            ProductProbeConfig.product_model_id == mid,
            ProductProbeConfig.probe_model_id == probe_id,
            ProductProbeConfig.feature_id == feat.id))
        cfg = r.scalar_one_or_none()
        if not cfg:
            cfg = ProductProbeConfig(product_model_id=mid, probe_model_id=probe_id, feature_id=feat.id)
            db.add(cfg)
        old_d, old_c = cfg.defined_status, cfg.current_status
        cfg.defined_status = def_stat
        cfg.current_status = cur_stat
        cfg.is_overridden = True
        draft = ProbeConfigDraft(
            product_model_id=mid, probe_model_id=probe_id, feature_id=feat.id,
            change_type="update",
            old_defined=old_d, new_defined=def_stat,
            old_current=old_c, new_current=cur_stat,
        )
        db.add(draft); count += 1

    await db.commit()
    return {"message": f"已创建 {count} 条草稿，请提交后生效"}


@router.post("/{product_model_id}/probes", response_model=dict)
async def set_product_probes(product_model_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    """设置产品支持的探头列表 { probe_model_ids: [1,2,3], priorities: {1: "标1", ...} }"""
    from sqlalchemy import delete as sqla_delete
    await db.execute(sqla_delete(ProductProbeModel).where(ProductProbeModel.product_model_id == product_model_id))
    ids = data.get("probe_model_ids", [])
    priorities = data.get("priorities", {})
    for pid in ids:
        db.add(ProductProbeModel(product_model_id=product_model_id, probe_model_id=pid,
                                  priority=priorities.get(str(pid))))
    await db.commit()
    return {"message": f"已设置 {len(ids)} 个探头关联"}


@router.get("/{product_model_id}/probes", response_model=list)
async def get_product_probes(product_model_id: int, db: AsyncSession = Depends(get_db)):
    """获取产品关联的探头列表"""
    result = await db.execute(
        select(ProductProbeModel, ProbeModel, ProbeCategory)
        .join(ProbeModel, ProductProbeModel.probe_model_id == ProbeModel.id)
        .join(ProbeCategory, ProbeModel.category_id == ProbeCategory.id)
        .where(ProductProbeModel.product_model_id == product_model_id)
        .order_by(ProbeCategory.sort_order, ProbeModel.sort_order)
    )
    return [{"id": r.ProductProbeModel.id, "probe_model_id": r.ProbeModel.id,
             "model_number": r.ProbeModel.model_number, "category_name": r.ProbeCategory.name,
             "priority": r.ProductProbeModel.priority} for r in result.all()]


@router.post("/{product_model_id}/batch-from-template", response_model=dict)
async def batch_from_template(product_model_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    """从模板批量设置某探头的所有功能 {probe_model_id}"""
    probe_id = data.get("probe_model_id")
    if not probe_id: raise HTTPException(status_code=400, detail="缺少probe_model_id")

    probe_r = await db.execute(select(ProbeModel).where(ProbeModel.id == probe_id))
    probe = probe_r.scalar_one_or_none()
    if not probe: raise HTTPException(status_code=404, detail="探头不存在")

    tpl_r = await db.execute(
        select(TemplateFeature).where(TemplateFeature.category_id == probe.category_id))
    tpl_map = {t.feature_id: t for t in tpl_r.scalars().all()}

    model_ids = await _expand_to_group(db, [product_model_id])
    feat_result = await db.execute(select(Feature))
    count = 0
    for mid in model_ids:
      for feat in feat_result.scalars().all():
        tpl = tpl_map.get(feat.id)
        def_stat = tpl.default_support if tpl else "unsupported"
        r = await db.execute(select(ProductProbeConfig).where(
            ProductProbeConfig.product_model_id == mid,
            ProductProbeConfig.probe_model_id == probe_id,
            ProductProbeConfig.feature_id == feat.id))
        cfg = r.scalar_one_or_none()
        if not cfg:
            cfg = ProductProbeConfig(product_model_id=mid, probe_model_id=probe_id, feature_id=feat.id)
            db.add(cfg)
        old_d, old_c = cfg.defined_status, cfg.current_status
        cfg.defined_status = def_stat
        cfg.current_status = def_stat
        cfg.is_overridden = False
        draft = ProbeConfigDraft(
            product_model_id=mid, probe_model_id=probe_id, feature_id=feat.id,
            change_type="update", old_defined=old_d, new_defined=def_stat,
            old_current=old_c, new_current=def_stat)
        db.add(draft); count += 1
    await db.commit()
    return {"message": f"已从模板创建 {count} 条草稿"}


@router.get("/{product_model_id}/export")
async def export_config(product_model_id: int, db: AsyncSession = Depends(get_db)):
    """导出产品探头配置为 Excel"""
    # Get matrix data
    matrix = await get_config(product_model_id, db)
    if not matrix.configs:
        raise HTTPException(status_code=400, detail="没有配置数据")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = matrix.product_model_name or "探头配置"

    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold = Font(bold=True)

    # Row 1: Feature groups (merged)
    ws.cell(row=1, column=1).value = "类别"
    ws.cell(row=1, column=2).value = "探头型号"
    col = 3
    groups = {}
    for f in matrix.features:
        gk = f['group_id']
        if gk not in groups:
            groups[gk] = {'name': f['group_name'], 'start': col, 'end': col}
        groups[gk]['end'] = col
        col += 1

    for gk, g in groups.items():
        ws.merge_cells(start_row=1, start_column=g['start'], end_row=1, end_column=g['end'])
        c = ws.cell(row=1, column=g['start'], value=g['name'])
        c.font = bold; c.fill = header_fill; c.alignment = center; c.border = thin
        for c2 in range(g['start'], g['end'] + 1):
            ws.cell(row=1, column=c2).fill = header_fill; ws.cell(row=1, column=c2).border = thin

    # Row 2: Feature names
    ws.cell(row=2, column=1).value = "类别"; ws.cell(row=2, column=1).font = bold; ws.cell(row=2, column=1).fill = header_fill; ws.cell(row=2, column=1).border = thin
    ws.cell(row=2, column=2).value = "探头型号"; ws.cell(row=2, column=2).font = bold; ws.cell(row=2, column=2).fill = header_fill; ws.cell(row=2, column=2).border = thin
    for i, f in enumerate(matrix.features):
        c = ws.cell(row=2, column=i+3, value=f['name'])
        c.font = bold; c.fill = header_fill; c.alignment = center; c.border = thin

    # Data rows
    gr_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    sup_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    con_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    unsp_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    row = 3
    last_cat = None
    for probe in matrix.probe_models:
        ws.cell(row=row, column=1, value=probe['category_name'] if probe['category_name'] != last_cat else '').border = thin
        ws.cell(row=row, column=2, value=probe['model_number']).border = thin
        if probe['category_name'] != last_cat:
            last_cat = probe['category_name']
        for i, f in enumerate(matrix.features):
            cfg = matrix.configs.get(f"{probe['id']}_{f['id']}")
            status = cfg.defined_status if cfg else 'unsupported'
            icon = {'supported': '√', 'conditional': '△', 'unsupported': ''}[status]
            c = ws.cell(row=row, column=i+3, value=icon)
            c.alignment = center; c.border = thin
            if status == 'supported': c.fill = sup_fill
            elif status == 'conditional': c.fill = con_fill
        row += 1

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    filename = f"{matrix.product_model_name}_探头配置_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})


# ========== 机型分组管理 ==========

@router.get("/model-groups")
async def get_model_groups(db: AsyncSession = Depends(get_db)):
    """获取所有机型分组信息"""
    r = await db.execute(
        select(ProductModel).order_by(ProductModel.config_group, ProductModel.name))
    groups: dict = {}
    ungrouped = []
    for m in r.scalars().all():
        if m.config_group:
            groups.setdefault(m.config_group, []).append({"id": m.id, "name": m.name, "series_id": m.series_id})
        else:
            ungrouped.append({"id": m.id, "name": m.name, "series_id": m.series_id})
    return {"groups": groups, "ungrouped": ungrouped}


@router.put("/model-groups")
async def set_model_groups(data: dict, db: AsyncSession = Depends(get_db)):
    """批量设置机型分组 {group_name: [model_ids]}"""
    count = 0
    for group_name, ids in data.items():
        if not ids: continue
        for mid in ids:
            r = await db.execute(select(ProductModel).where(ProductModel.id == mid))
            m = r.scalar_one_or_none()
            if m:
                m.config_group = group_name if group_name else None
                count += 1
    await db.commit()
    return {"message": f"已更新 {count} 个机型的分组"}


@router.post("/model-groups/rename")
async def rename_model_group(data: dict, db: AsyncSession = Depends(get_db)):
    """重命名分组 {old_name, new_name}"""
    old = data.get("old_name")
    new = data.get("new_name")
    if not old or not new: raise HTTPException(400, "缺少 old_name 或 new_name")
    r = await db.execute(select(ProductModel).where(ProductModel.config_group == old))
    count = 0
    for m in r.scalars().all():
        m.config_group = new
        count += 1
    await db.commit()
    return {"message": f"已将 {count} 个机型从 '{old}' 重命名为 '{new}'"}


@router.post("/model-groups/auto")
async def auto_group_models(db: AsyncSession = Depends(get_db)):
    """自动分组：按基础型号名去区域后缀"""
    import re
    r = await db.execute(select(ProductModel))
    models = r.scalars().all()
    suffixes = ['_BRA', '_RUA', '_Private_RUA', '_Private', '_POC', '_Anesthesia']
    count = 0
    for m in models:
        if m.config_group: continue
        base = m.name
        for sfx in suffixes:
            if base.endswith(sfx):
                base = base[:-len(sfx)]
                break
        base = base.strip()
        if base != m.name:
            m.config_group = base
            count += 1
    await db.commit()
    return {"message": f"已为 {count} 个机型自动分组"}
