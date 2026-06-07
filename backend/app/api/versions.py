"""
版本管理 API
"""
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func
from sqlalchemy.orm import selectinload
from typing import List, Optional
import json
import uuid

from app.database import get_db
from app.models import ConfigVersion, ProductSeries, ConfigItem, ConfigValue, ProductModel, ChangeLog
from app.schemas.version import (
    ConfigVersionCreate, ConfigVersionResponse,
    VersionCompareRequest, VersionCompareResponse, VersionDiffDetail,
    ConfigVersionListResponse
)
from app.utils import generate_next_version

router = APIRouter()


@router.get("", response_model=ConfigVersionListResponse)
async def get_versions(
    series_id: int = None,
    skip: int = 0,
    limit: int = 20,
    db: AsyncSession = Depends(get_db)
):
    """获取版本列表"""
    query = select(ConfigVersion)

    if series_id:
        query = query.where(ConfigVersion.series_id == series_id)

    query = query.order_by(ConfigVersion.id.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    items = result.scalars().all()

    return ConfigVersionListResponse(items=items, total=len(items))


# ==================== 变更日志 API ====================
# 注意：静态路由必须在动态路由（/{version_id}）之前定义

@router.get("/change-logs")
async def get_change_logs(
    series_id: int = None,
    version_id: int = None,
    change_type: str = None,
    skip: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_db)
):
    """
    获取变更日志列表

    Args:
        series_id: 按产品系列筛选
        version_id: 按版本筛选
        change_type: 按变更类型筛选 (create/update/delete)
        skip: 分页偏移
        limit: 分页限制
    """
    query = select(ChangeLog)

    if series_id:
        query = query.where(ChangeLog.series_id == series_id)
    if version_id:
        query = query.where(ChangeLog.version_id == version_id)
    if change_type:
        query = query.where(ChangeLog.change_type == change_type)

    query = query.order_by(ChangeLog.changed_at.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    logs = result.scalars().all()

    # 批量查询关联信息，避免 N+1 问题
    item_ids = [log.item_id for log in logs if log.item_id]
    model_ids = [log.model_id for log in logs if log.model_id]

    items_map = {}
    if item_ids:
        items_result = await db.execute(
            select(ConfigItem).where(ConfigItem.id.in_(item_ids))
        )
        items_map = {i.id: i for i in items_result.scalars().all()}

    models_map = {}
    if model_ids:
        models_result = await db.execute(
            select(ProductModel).where(ProductModel.id.in_(model_ids))
        )
        models_map = {m.id: m for m in models_result.scalars().all()}

    # 获取关联的配置项和型号信息
    logs_with_details = []
    for log in logs:
        log_data = {
            "id": log.id,
            "series_id": log.series_id,
            "version_id": log.version_id,
            "change_type": log.change_type,
            "item_id": log.item_id,
            "model_id": log.model_id,
            "field_name": log.field_name,
            "old_value": log.old_value,
            "new_value": log.new_value,
            "changed_by": log.changed_by,
            "changed_at": log.changed_at.isoformat() if log.changed_at else None
        }

        # 从批量查询结果中获取配置项信息
        if log.item_id:
            item = items_map.get(log.item_id)
            if item:
                log_data["item_rd_name"] = item.rd_name
                log_data["item_ipn"] = item.ipn

        # 从批量查询结果中获取型号信息
        if log.model_id:
            model = models_map.get(log.model_id)
            if model:
                log_data["model_name"] = model.name

        logs_with_details.append(log_data)

    # 获取总数
    count_query = select(func.count()).select_from(ChangeLog)
    if series_id:
        count_query = count_query.where(ChangeLog.series_id == series_id)
    if version_id:
        count_query = count_query.where(ChangeLog.version_id == version_id)
    if change_type:
        count_query = count_query.where(ChangeLog.change_type == change_type)

    count_result = await db.execute(count_query)
    total = count_result.scalar()

    return {
        "items": logs_with_details,
        "total": total
    }


@router.get("/change-logs/{log_id}")
async def get_change_log(
    log_id: int,
    db: AsyncSession = Depends(get_db)
):
    """获取单个变更日志详情"""
    result = await db.execute(select(ChangeLog).where(ChangeLog.id == log_id))
    log = result.scalar_one_or_none()

    if not log:
        raise HTTPException(status_code=404, detail="变更记录不存在")

    log_data = {
        "id": log.id,
        "series_id": log.series_id,
        "version_id": log.version_id,
        "change_type": log.change_type,
        "item_id": log.item_id,
        "model_id": log.model_id,
        "field_name": log.field_name,
        "old_value": log.old_value,
        "new_value": log.new_value,
        "changed_by": log.changed_by,
        "changed_at": log.changed_at.isoformat() if log.changed_at else None
    }

    # 获取配置项信息
    if log.item_id:
        item_result = await db.execute(
            select(ConfigItem).where(ConfigItem.id == log.item_id)
        )
        item = item_result.scalar_one_or_none()
        if item:
            log_data["item"] = {
                "id": item.id,
                "rd_name": item.rd_name,
                "ipn": item.ipn,
                "category": item.category
            }

    # 获取型号信息
    if log.model_id:
        model_result = await db.execute(
            select(ProductModel).where(ProductModel.id == log.model_id)
        )
        model = model_result.scalar_one_or_none()
        if model:
            log_data["model"] = {
                "id": model.id,
                "name": model.name
            }

    # 获取版本信息
    if log.version_id:
        version_result = await db.execute(
            select(ConfigVersion).where(ConfigVersion.id == log.version_id)
        )
        version = version_result.scalar_one_or_none()
        if version:
            log_data["version"] = {
                "id": version.id,
                "version_number": version.version_number,
                "version_name": version.version_name
            }

    return log_data


# ==================== 版本详情 API ====================

@router.get("/{version_id}", response_model=ConfigVersionResponse)
async def get_version(
    version_id: int,
    db: AsyncSession = Depends(get_db)
):
    """获取版本详情"""
    result = await db.execute(select(ConfigVersion).where(ConfigVersion.id == version_id))
    version = result.scalar_one_or_none()

    if not version:
        raise HTTPException(status_code=404, detail="版本不存在")

    return version


@router.post("", response_model=ConfigVersionResponse)
async def create_version(
    data: ConfigVersionCreate,
    db: AsyncSession = Depends(get_db)
):
    """创建新版本（从当前数据快照）"""
    # 检查系列是否存在
    result = await db.execute(select(ProductSeries).where(ProductSeries.id == data.series_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="产品系列不存在")

    # 获取最新版本号
    last_version_result = await db.execute(
        select(ConfigVersion)
        .where(ConfigVersion.series_id == data.series_id)
        .order_by(ConfigVersion.id.desc())
        .limit(1)
    )
    last_version = last_version_result.scalar_one_or_none()

    # 生成版本号
    if data.version_number:
        version_number = data.version_number
    else:
        if last_version:
            version_number = generate_next_version(last_version.version_number)
        else:
            version_number = "v1.0.0"

    # 获取当前数据快照
    models_result = await db.execute(
        select(ProductModel).where(ProductModel.series_id == data.series_id)
    )
    models = models_result.scalars().all()
    model_ids = [m.id for m in models]

    items_result = await db.execute(select(ConfigItem).order_by(ConfigItem.row_index))
    items = items_result.scalars().all()

    values_result = await db.execute(
        select(ConfigValue).where(ConfigValue.model_id.in_(model_ids))
    )
    values = values_result.scalars().all()

    # 构建快照数据
    value_map = {}
    for v in values:
        if v.item_id not in value_map:
            value_map[v.item_id] = {}
        value_map[v.item_id][v.model_id] = {
            "current_config": v.current_config,
            "final_config": v.final_config,
            "selection_config": v.selection_config,
            "rd_status": v.rd_status
        }

    # 如果已有版本，对比数据是否有变化
    if last_version and last_version.snapshot_data:
        try:
            prev_snapshot = json.loads(last_version.snapshot_data)
            prev_value_map = {}
            for item in prev_snapshot.get("items", []):
                iid = item["id"]
                for mid_str, vals in item.get("values", {}).items():
                    mid = int(mid_str)
                    if iid not in prev_value_map:
                        prev_value_map[iid] = {}
                    prev_value_map[iid][mid] = vals

            if prev_value_map == value_map:
                raise HTTPException(
                    status_code=400,
                    detail="当前数据与上次版本无任何变化，无需创建新版本"
                )
        except HTTPException:
            raise
        except Exception:
            pass  # 快照解析失败则跳过对比

    snapshot = {
        "models": [{"id": m.id, "name": m.name} for m in models],
        "items": [
            {
                "id": item.id,
                "category": item.category,
                "row_index": item.row_index,
                "rd_name": item.rd_name,
                "v_code": item.v_code,
                "ipn": item.ipn,
                "zh_desc": item.zh_desc,
                "en_desc": item.en_desc,
                "values": value_map.get(item.id, {})
            }
            for item in items
        ]
    }

    # 创建版本
    version = ConfigVersion(
        series_id=data.series_id,
        version_number=version_number,
        version_name=data.version_name,
        description=data.description,
        snapshot_data=json.dumps(snapshot, ensure_ascii=False),
        row_count=len(items)
    )

    db.add(version)
    await db.commit()
    await db.refresh(version)

    return version


@router.post("/compare", response_model=VersionCompareResponse)
async def compare_versions(
    data: VersionCompareRequest,
    db: AsyncSession = Depends(get_db)
):
    """对比两个版本"""
    # 获取两个版本
    result1 = await db.execute(select(ConfigVersion).where(ConfigVersion.id == data.version_id_1))
    version1 = result1.scalar_one_or_none()

    result2 = await db.execute(select(ConfigVersion).where(ConfigVersion.id == data.version_id_2))
    version2 = result2.scalar_one_or_none()

    if not version1 or not version2:
        raise HTTPException(status_code=404, detail="版本不存在")

    # 解析快照
    snapshot1 = json.loads(version1.snapshot_data)
    snapshot2 = json.loads(version2.snapshot_data)

    # 构建索引
    def build_index(snapshot):
        index = {}
        for item in snapshot.get("items", []):
            key = (item.get("row_index"), item.get("ipn"))
            index[key] = item
        return index

    index1 = build_index(snapshot1)
    index2 = build_index(snapshot2)

    # 计算差异
    added = []
    modified = []
    deleted = []

    keys1 = set(index1.keys())
    keys2 = set(index2.keys())

    # 新增项
    for key in keys2 - keys1:
        item = index2[key]
        added.append({
            "type": "added",
            "row_index": item.get("row_index"),
            "rd_name": item.get("rd_name"),
            "ipn": item.get("ipn")
        })

    # 删除项
    for key in keys1 - keys2:
        item = index1[key]
        deleted.append({
            "type": "deleted",
            "row_index": item.get("row_index"),
            "rd_name": item.get("rd_name"),
            "ipn": item.get("ipn")
        })

    # 修改项
    for key in keys1 & keys2:
        item1 = index1[key]
        item2 = index2[key]

        values1 = item1.get("values", {})
        values2 = item2.get("values", {})

        # 如果指定了机型筛选，只对比指定的机型
        model_ids_to_check = data.model_ids if data.model_ids else list(set(values1.keys()) | set(values2.keys()))

        for model_id in model_ids_to_check:
            model_id_str = str(model_id)
            v1 = values1.get(model_id_str, {})
            v2 = values2.get(model_id_str, {})

            for field in ["current_config", "final_config", "selection_config", "rd_status"]:
                old = v1.get(field)
                new = v2.get(field)

                # 忽略空值和N/A的差异
                old_normalized = old if old not in [None, "", "N/A"] else None
                new_normalized = new if new not in [None, "", "N/A"] else None

                if old_normalized != new_normalized:
                    modified.append({
                        "type": "modified",
                        "row_index": item1.get("row_index"),
                        "rd_name": item1.get("rd_name"),
                        "ipn": item1.get("ipn"),
                        "model_id": int(model_id),
                        "field_name": field,
                        "old_value": old,
                        "new_value": new
                    })

    return {
        "version_1": version1,
        "version_2": version2,
        "added": added,
        "modified": modified,
        "deleted": deleted,
        "summary": {
            "added": len(added),
            "modified": len(modified),
            "deleted": len(deleted)
        }
    }


@router.delete("/{version_id}")
async def delete_version(
    version_id: int,
    db: AsyncSession = Depends(get_db)
):
    """删除版本"""
    result = await db.execute(select(ConfigVersion).where(ConfigVersion.id == version_id))
    version = result.scalar_one_or_none()

    if not version:
        raise HTTPException(status_code=404, detail="版本不存在")

    await db.delete(version)
    await db.commit()
    return {"message": "删除成功"}


@router.put("/{version_id}", response_model=ConfigVersionResponse)
async def update_version(
    version_id: int,
    version_number: Optional[str] = None,
    version_name: Optional[str] = None,
    description: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """更新版本信息（版本号、名称、说明）"""
    result = await db.execute(select(ConfigVersion).where(ConfigVersion.id == version_id))
    version = result.scalar_one_or_none()

    if not version:
        raise HTTPException(status_code=404, detail="版本不存在")

    # 检查版本号是否重复
    if version_number and version_number != version.version_number:
        existing = await db.execute(
            select(ConfigVersion).where(
                ConfigVersion.series_id == version.series_id,
                ConfigVersion.version_number == version_number
            )
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="版本号已存在")
        version.version_number = version_number

    if version_name is not None:
        version.version_name = version_name
    if description is not None:
        version.description = description

    await db.commit()
    await db.refresh(version)
    return version


@router.post("/{version_id}/rollback")
async def rollback_version(
    version_id: int,
    db: AsyncSession = Depends(get_db)
):
    """回滚到指定版本（创建新版本，内容为目标版本快照）"""
    result = await db.execute(select(ConfigVersion).where(ConfigVersion.id == version_id))
    target_version = result.scalar_one_or_none()

    if not target_version:
        raise HTTPException(status_code=404, detail="目标版本不存在")

    # 解析目标版本快照
    snapshot = json.loads(target_version.snapshot_data)

    # 获取当前型号
    models_result = await db.execute(
        select(ProductModel).where(ProductModel.series_id == target_version.series_id)
    )
    current_models = {m.id: m for m in models_result.scalars().all()}

    # 清空当前配置数据
    await db.execute(delete(ConfigValue).where(ConfigValue.model_id.in_(current_models.keys())))
    await db.execute(delete(ConfigItem))

    # 恢复配置项和配置值
    model_id_map = {}  # 快照中的model_id -> 当前model_id
    for model_info in snapshot.get("models", []):
        # 查找或创建型号
        model = await db.execute(
            select(ProductModel).where(
                ProductModel.series_id == target_version.series_id,
                ProductModel.name == model_info.get("name")
            )
        )
        existing_model = model.scalar_one_or_none()
        if existing_model:
            model_id_map[model_info.get("id")] = existing_model.id

    # 恢复配置项和值
    new_items = []
    for item_info in snapshot.get("items", []):
        item = ConfigItem(
            category=item_info.get("category"),
            row_index=item_info.get("row_index"),
            rd_name=item_info.get("rd_name"),
            v_code=item_info.get("v_code"),
            ipn=item_info.get("ipn"),
            zh_desc=item_info.get("zh_desc"),
            en_desc=item_info.get("en_desc")
        )
        db.add(item)
        await db.flush()
        new_items.append((item, item_info))

    # 恢复配置值
    for item, item_info in new_items:
        values = item_info.get("values", {})
        for old_model_id, value_info in values.items():
            new_model_id = model_id_map.get(int(old_model_id))
            if new_model_id:
                value = ConfigValue(
                    item_id=item.id,
                    model_id=new_model_id,
                    current_config=value_info.get("current_config"),
                    final_config=value_info.get("final_config"),
                    selection_config=value_info.get("selection_config"),
                    rd_status=value_info.get("rd_status")
                )
                db.add(value)

    # 创建新版本
    last_version = await db.execute(
        select(ConfigVersion)
        .where(ConfigVersion.series_id == target_version.series_id)
        .order_by(ConfigVersion.id.desc())
        .limit(1)
    )
    last = last_version.scalar_one_or_none()

    new_version_number = generate_next_version(last.version_number if last else None)

    new_version = ConfigVersion(
        series_id=target_version.series_id,
        version_number=new_version_number,
        version_name=f"回滚自 {target_version.version_number}",
        description=f"回滚自版本 {target_version.version_number}",
        snapshot_data=target_version.snapshot_data,
        row_count=len(new_items)
    )
    db.add(new_version)

    await db.commit()
    await db.refresh(new_version)

    return {
        "message": "回滚成功",
        "new_version": new_version
    }


@router.post("/compare/export")
async def export_version_compare(
    data: VersionCompareRequest,
    db: AsyncSession = Depends(get_db)
):
    """导出版本对比结果为Excel"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    # 获取对比结果
    result1 = await db.execute(select(ConfigVersion).where(ConfigVersion.id == data.version_id_1))
    version1 = result1.scalar_one_or_none()

    result2 = await db.execute(select(ConfigVersion).where(ConfigVersion.id == data.version_id_2))
    version2 = result2.scalar_one_or_none()

    if not version1 or not version2:
        raise HTTPException(status_code=404, detail="版本不存在")

    # 解析快照
    snapshot1 = json.loads(version1.snapshot_data)
    snapshot2 = json.loads(version2.snapshot_data)

    # 构建索引
    def build_index(snapshot):
        index = {}
        for item in snapshot.get("items", []):
            key = (item.get("row_index"), item.get("ipn"))
            index[key] = item
        return index

    index1 = build_index(snapshot1)
    index2 = build_index(snapshot2)

    # 计算差异
    added = []
    modified = []
    deleted = []

    keys1 = set(index1.keys())
    keys2 = set(index2.keys())

    for key in keys2 - keys1:
        item = index2[key]
        added.append(item)

    for key in keys1 - keys2:
        item = index1[key]
        deleted.append(item)

    for key in keys1 & keys2:
        item1 = index1[key]
        item2 = index2[key]
        values1 = item1.get("values", {})
        values2 = item2.get("values", {})

        for model_id in set(values1.keys()) | set(values2.keys()):
            v1 = values1.get(model_id, {})
            v2 = values2.get(model_id, {})

            for field in ["current_config", "final_config", "selection_config", "rd_status"]:
                old = v1.get(field)
                new = v2.get(field)
                old_normalized = old if old not in [None, "", "N/A"] else None
                new_normalized = new if new not in [None, "", "N/A"] else None

                if old_normalized != new_normalized:
                    model_name = ""
                    for m in snapshot2.get("models", []):
                        if str(m.get("id")) == str(model_id):
                            model_name = m.get("name", "")
                            break

                    modified.append({
                        "rd_name": item1.get("rd_name"),
                        "ipn": item1.get("ipn"),
                        "model_name": model_name,
                        "field": field,
                        "old_value": old or "",
                        "new_value": new or ""
                    })

    # 创建Excel
    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True)
    added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    deleted_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    modified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 新增项sheet
    ws1 = wb.active
    ws1.title = "新增项"
    ws1.append(["研发名称", "IPN号", "分类"])
    for cell in ws1[1]:
        cell.font = header_font
    for item in added:
        ws1.append([item.get("rd_name"), item.get("ipn"), item.get("category")])
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            cell.fill = added_fill
            cell.border = thin_border

    # 删除项sheet
    ws2 = wb.create_sheet("删除项")
    ws2.append(["研发名称", "IPN号", "分类"])
    for cell in ws2[1]:
        cell.font = header_font
    for item in deleted:
        ws2.append([item.get("rd_name"), item.get("ipn"), item.get("category")])
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.fill = deleted_fill
            cell.border = thin_border

    # 修改项sheet
    ws3 = wb.create_sheet("修改项")
    ws3.append(["研发名称", "IPN号", "型号", "字段", "原值", "新值"])
    for cell in ws3[1]:
        cell.font = header_font
    field_labels = {
        "final_config": "最终配置",
        "current_config": "当前配置",
        "selection_config": "选型类别",
        "rd_status": "研发状态"
    }
    for item in modified:
        ws3.append([
            item.get("rd_name"),
            item.get("ipn"),
            item.get("model_name"),
            field_labels.get(item.get("field"), item.get("field")),
            item.get("old_value"),
            item.get("new_value")
        ])
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            cell.fill = modified_fill
            cell.border = thin_border

    # 设置列宽
    for ws in [ws1, ws2, ws3]:
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        if ws == ws3:
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"version_compare_{version1.version_number}_{version2.version_number}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )