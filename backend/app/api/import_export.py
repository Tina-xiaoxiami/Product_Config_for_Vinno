"""
导入导出 API
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func
from typing import List, Optional
from pydantic import BaseModel
import io
import json


class ExportRequest(BaseModel):
    """导出请求参数"""
    series_id: int
    include_main_unit: bool = False
    item_ids: Optional[str] = None
    categories: Optional[str] = None
    search: Optional[str] = None
    model_ids: Optional[str] = None
    visible_fields: Optional[str] = None
    draft_changes: Optional[str] = None
    deleted_items: Optional[str] = None
    new_items: Optional[str] = None
import uuid
from datetime import datetime

from app.database import get_db
from app.models import (
    ProductSeries, ProductModel, ConfigItem, ConfigValue,
    DraftBatch, ConfigDraft, ImportHistory, ConfigVersion
)
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter()


def parse_merged_cells(ws):
    """解析合并单元格信息，返回每个合并区域的起始单元格值"""
    merged_info = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
        value = ws.cell(row=min_row, column=min_col).value
        merged_info[(min_row, min_col)] = {
            'value': value,
            'max_col': max_col,
            'max_row': max_row
        }
    return merged_info


@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    series_name: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """
    导入Excel文件

    Excel结构：
    - 第1行：产品系列（合并单元格）
    - 第2行：产品型号（合并单元格，每个型号跨4列）
    - 第3行：配置状态（最终配置、当前配置、选型类别、研发状态）
    - 第4行：分类标题行（如 Main Unit）
    - 第5行开始：配置数据行

    A-E列固定：研发名称、V代码、IPN号、中文描述、英文描述
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持Excel文件(.xlsx, .xls)")

    # 读取文件
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # 解析合并单元格
    merged_info = parse_merged_cells(ws)

    # 解析产品系列（第1行，从F列开始）
    # 收集所有合并区域，按连续范围分别存储（不合并同名系列）
    series_ranges_list = []  # [(name, min_col, max_col), ...]
    processed_cols = set()

    for col in range(6, ws.max_column + 1):
        if col in processed_cols:
            continue
        if (1, col) in merged_info:
            info = merged_info[(1, col)]
            series_name_from_excel = str(info['value']).strip() if info['value'] else None
            if series_name_from_excel:
                # 标记已处理的列
                for c in range(col, info['max_col'] + 1):
                    processed_cols.add(c)
                # 不合并同名系列，每个连续区域单独存储
                series_ranges_list.append((series_name_from_excel, col, info['max_col']))

    # 合并相邻的同名系列范围（同一系列在 Excel 中可能是每 4 列一个连续合并区域）
    # 注意：只合并相邻/重叠的范围，不合并有间隔的（间隔中可能包含其他系列）
    sorted_ranges = sorted(series_ranges_list, key=lambda x: x[1])  # 按 col_start 排序
    merged_series = []  # [(name, col_start, col_end)]
    for name, col_start, col_end in sorted_ranges:
        if merged_series and merged_series[-1][0] == name and merged_series[-1][2] + 1 >= col_start:
            # 同名且相邻/重叠 → 合并范围
            merged_series[-1] = (name, merged_series[-1][1], max(merged_series[-1][2], col_end))
        else:
            merged_series.append([name, col_start, col_end])

    # 同名系列可能有多个非连续范围（如 VINNO 9_Private 远离其他型号列）
    # 合并为一条记录，存储所有子范围，避免同系列被多次迭代导致变更检测错误
    series_groups = {}  # name -> [(col_start, col_end), ...]
    for name, col_start, col_end in merged_series:
        if name not in series_groups:
            series_groups[name] = []
        series_groups[name].append((col_start, col_end))

    series_list = []
    for name, ranges in series_groups.items():
        series_list.append({
            'name': name,
            'ranges': ranges,  # 所有子范围列表，模型解析时逐个迭代
            'col_start': min(r[0] for r in ranges),  # 最早列
            'col_end': max(r[1] for r in ranges),    # 最晚列
        })

    # 如果没有解析到系列，使用文件名
    if not series_list:
        series_name = series_name or file.filename.replace('.xlsx', '').replace('.xls', '')
        series_list.append({
            'name': series_name,
            'col_start': 6,
            'col_end': ws.max_column
        })

    results = []
    change_log = []  # 变更记录

    try:
        for series_info in series_list:
            current_series_name = series_info['name']
            col_start = series_info['col_start']
            col_end = series_info['col_end']

            # 创建或获取产品系列
            series_result = await db.execute(
                select(ProductSeries).where(ProductSeries.name == current_series_name)
            )
            series = series_result.scalar_one_or_none()

            if not series:
                series = ProductSeries(name=current_series_name)
                db.add(series)
                await db.flush()

            # 获取最后发布的版本快照（用于对照变更）
            last_version_result = await db.execute(
                select(ConfigVersion)
                .where(ConfigVersion.series_id == series.id)
                .order_by(ConfigVersion.id.desc())
                .limit(1)
            )
            last_version = last_version_result.scalar_one_or_none()
            # 用IPN索引快照数据（回滚会重建ConfigItem导致ID变化，IPN是稳定标识）
            snapshot_raw_data = None
            # 预计算快照值字典：{(ipn_str, 型号名称, 字段名): value}，O(1) 查询
            snapshot_values = {}
            # 快照中的 (IPN, 型号名称) 对集合，用于判定新增/删除/修改
            snapshot_pairs = set()
            if last_version and last_version.snapshot_data:
                snapshot_raw_data = json.loads(last_version.snapshot_data)
                # 构建快照中型号ID→名称映射（回滚后model_id也会变，需用名称匹配）
                snapshot_model_name_map = {}
                for m in snapshot_raw_data.get("models", []):
                    mid = m.get("id")
                    if mid:
                        snapshot_model_name_map[mid] = m.get("name")

                for item_entry in snapshot_raw_data.get("items", []):
                    # 与 Excel 解析保持一致，跳过 Main Unit 类别的项
                    if item_entry.get("category") == "Main Unit":
                        continue
                    ipn = item_entry.get("ipn")
                    if not ipn:
                        continue
                    ipn_str = str(ipn).strip()
                    for snap_model_id_str, model_vals in item_entry.get("values", {}).items():
                        snap_model_name = snapshot_model_name_map.get(int(snap_model_id_str))
                        if not snap_model_name:
                            continue
                        snapshot_pairs.add((ipn_str, snap_model_name))
                        for field in ("final_config", "current_config", "selection_config", "rd_status"):
                            val = model_vals.get(field) if model_vals else None
                            snapshot_values[(ipn_str, snap_model_name, field)] = val

            # 计算快照中所有字段均为 N/A 值的配对 —— 这些应视为"新增"而非"修改"
            NA_VALUES = {'N/A', '', None, '-', 'None', 'null', '未定义'}
            snapshot_all_na_pairs = set()
            for ipn_str, snap_model_name in snapshot_pairs:
                all_na = all(
                    snapshot_values.get((ipn_str, snap_model_name, f)) in NA_VALUES
                    for f in ("final_config", "current_config", "selection_config", "rd_status")
                )
                if all_na:
                    snapshot_all_na_pairs.add((ipn_str, snap_model_name))

            # 预加载所有已存在的型号（跨系列），避免重复创建
            existing_models_map = {}
            all_models_result = await db.execute(select(ProductModel))
            for m in all_models_result.scalars().all():
                if m.name not in existing_models_map:
                    existing_models_map[m.name] = []
                existing_models_map[m.name].append(m)

            # 解析产品型号（第2行），遍历该系列的所有子列范围
            models = []
            for range_start, range_end in series_info.get('ranges', [(col_start, col_end)]):
                col = range_start
                while col <= range_end:
                    cell = ws.cell(row=2, column=col)
                    model_name = None

                    # 检查是否是合并单元格的起始
                    if (2, col) in merged_info:
                        info = merged_info[(2, col)]
                        raw_value = info['value']
                        model_end_col = info['max_col']
                    else:
                        raw_value = cell.value
                        model_end_col = col + 3  # 默认4列

                    if raw_value:
                        # 处理格式：型号名//uuid
                        model_name = str(raw_value).split('//')[0].strip()

                        # 检查当前系列内是否已有该型号
                        model_result = await db.execute(
                            select(ProductModel).where(
                                ProductModel.series_id == series.id,
                                ProductModel.name == model_name
                            )
                        )
                        model = model_result.scalar_one_or_none()

                        if not model:
                            # 检查其他系列是否已有同名型号
                            if model_name in existing_models_map:
                                existing_series = [m.series_id for m in existing_models_map[model_name]]
                                print(f"  注意: 型号 {model_name} 在系列 {existing_series} 中已存在，将在当前系列创建新实例")

                            model = ProductModel(
                                series_id=series.id,
                                name=model_name,
                                column_start=col,
                                column_end=model_end_col,
                                sort_order=len(models)
                            )
                            db.add(model)
                            await db.flush()

                            # 更新预加载的map
                            if model_name not in existing_models_map:
                                existing_models_map[model_name] = []
                            existing_models_map[model_name].append(model)
                        else:
                            # 重用已有型号时，更新列范围以匹配当前 Excel 中的实际位置
                            model.column_start = col
                            model.column_end = model_end_col

                        models.append(model)
                        col = model_end_col + 1
                    else:
                        col += 1

            # 解析配置数据（从第5行开始）
            current_category = None
            items_created = 0
            values_created = 0

            # 批量收集待创建的配置项和配置值
            items_to_create = []
            all_ipns = set()

            for row_idx in range(4, ws.max_row + 1):
                first_col = ws.cell(row=row_idx, column=1).value

                # 检查是否是分类标题行
                if first_col and isinstance(first_col, str):
                    stripped = first_col.strip()
                    # 识别分类标题（6种分类）
                    valid_categories = [
                        "Main Unit",
                        "Optional Features",
                        "Optional peripherals",
                        "*Optional peripherals(Preassemble in Factory)",
                        "Probes",
                        "Biopsy guide"
                    ]
                    if stripped in valid_categories or stripped.startswith("Optional"):
                        current_category = stripped
                        continue

                # 跳过Main Unit分类的数据
                if current_category == "Main Unit":
                    continue

                # 解析固定列（A-E列）
                rd_name = ws.cell(row=row_idx, column=1).value
                v_code = ws.cell(row=row_idx, column=2).value
                ipn = ws.cell(row=row_idx, column=3).value
                zh_desc = ws.cell(row=row_idx, column=4).value
                en_desc = ws.cell(row=row_idx, column=5).value

                # 跳过空行
                if not rd_name and not ipn:
                    continue

                ipn_str = str(ipn).strip() if ipn else None
                if ipn_str:
                    all_ipns.add(ipn_str)

                # 收集配置项数据（待批量创建）
                items_to_create.append({
                    'row_idx': row_idx,
                    'category': current_category or "Optional Features",
                    'rd_name': str(rd_name).strip() if rd_name else None,
                    'v_code': str(v_code).strip() if v_code else None,
                    'ipn': ipn_str,
                    'zh_desc': str(zh_desc).strip() if zh_desc else None,
                    'en_desc': str(en_desc).strip() if en_desc else None,
                })

            # 批量查询已存在的 ConfigItem（按 IPN）
            existing_items_map = {}
            if all_ipns:
                existing_items_result = await db.execute(
                    select(ConfigItem).where(ConfigItem.ipn.in_(all_ipns))
                )
                for item in existing_items_result.scalars().all():
                    existing_items_map[item.ipn] = item

            # 批量创建配置项
            created_items = []
            changes = []  # 记录该系列的变更
            draft_changes = []  # 记录草稿级变更（用于创建 ConfigDraft）
            processed_ipns = {}  # 跟踪本次导入已处理的 IPN，避免同文件内重复

            for item_data in items_to_create:
                ipn_str = item_data['ipn']
                change_type = None

                # 检查是否已存在（数据库中或本次导入中）
                if ipn_str and (ipn_str in existing_items_map or ipn_str in processed_ipns):
                    # 优先使用数据库中已存在的，否则使用本次导入创建的
                    if ipn_str in existing_items_map:
                        item = existing_items_map[ipn_str]
                    else:
                        item = processed_ipns[ipn_str]
                        # 跳过重复项的处理，直接添加配置值
                        created_items.append({"item": item, "data": item_data, "change_type": change_type})
                        continue

                    # 检查是否有字段变化
                    has_changes = (
                        item.rd_name != item_data['rd_name'] or
                        item.v_code != item_data['v_code'] or
                        item.zh_desc != item_data['zh_desc'] or
                        item.en_desc != item_data['en_desc'] or
                        item.category != item_data['category']
                    )
                    if has_changes:
                        change_type = "update"
                        changes.append({
                            "type": "update",
                            "item_id": item.id,
                            "ipn": ipn_str,
                            "rd_name": item_data['rd_name'],
                            "message": f"更新配置项: {item_data['rd_name']} ({ipn_str})"
                        })
                    # 更新现有记录的字段
                    item.rd_name = item_data['rd_name']
                    item.v_code = item_data['v_code']
                    item.zh_desc = item_data['zh_desc']
                    item.en_desc = item_data['en_desc']
                    item.row_index = item_data['row_idx']
                    item.category = item_data['category']
                else:
                    # 创建新的配置项
                    change_type = "create"
                    item = ConfigItem(
                        category=item_data['category'],
                        row_index=item_data['row_idx'],
                        rd_name=item_data['rd_name'],
                        v_code=item_data['v_code'],
                        ipn=item_data['ipn'],
                        zh_desc=item_data['zh_desc'],
                        en_desc=item_data['en_desc']
                    )
                    db.add(item)
                    # 记录到已处理字典中
                    if ipn_str:
                        existing_items_map[ipn_str] = item
                        processed_ipns[ipn_str] = item

                created_items.append({"item": item, "data": item_data, "change_type": change_type})
                items_created += 1

            # 批量 flush 以获取所有 item.id
            await db.flush()

            # 将新建的配置项记录到 changes 中（用于后续创建草稿）
            for item_info in created_items:
                if item_info["change_type"] == "create" and item_info["item"].id:
                    changes.append({
                        "type": "create",
                        "item_id": item_info["item"].id,
                        "ipn": item_info["data"].get("ipn"),
                        "rd_name": item_info["data"].get("rd_name"),
                        "message": f"新建配置项: {item_info['data'].get('rd_name')} ({item_info['data'].get('ipn') or '无IPN'})"
                    })

            # 批量查询已存在的 ConfigValue（按 item_id + model_id）
            item_ids = [item_info["item"].id for item_info in created_items if item_info["item"].id]
            model_ids = [m.id for m in models if m.id]
            existing_values_map = {}

            # 首先查询数据库中已存在的配置值
            if item_ids and model_ids:
                existing_values_result = await db.execute(
                    select(ConfigValue).where(
                        ConfigValue.item_id.in_(item_ids),
                        ConfigValue.model_id.in_(model_ids)
                    )
                )
                for val in existing_values_result.scalars().all():
                    existing_values_map[(val.item_id, val.model_id)] = val

            # 批量创建配置值 - 使用字典避免重复
            values_to_create = {}  # key: (item_id, model_id), value: ConfigValue
            excel_pairs = set()  # 跟踪Excel中所有 (IPN, 型号名) 对

            for item_info in created_items:
                item = item_info["item"]
                item_data = item_info["data"]
                item_ipn = str(item.ipn).strip() if item and item.ipn else None

                for model in models:
                    model_col = model.column_start

                    # 读取4个配置状态
                    final_config = ws.cell(row=item_data['row_idx'], column=model_col).value
                    current_config = ws.cell(row=item_data['row_idx'], column=model_col + 1).value
                    selection_config = ws.cell(row=item_data['row_idx'], column=model_col + 2).value
                    rd_status = ws.cell(row=item_data['row_idx'], column=model_col + 3).value

                    # 追踪Excel中的 (IPN, 型号名) 对
                    # 仅当4个字段至少有一个有意义值时加入配对，全N/A值不产生变更
                    pair_key = (item_ipn, model.name) if item_ipn else None
                    if pair_key:
                        excel_values = [
                            str(final_config).strip() if final_config else None,
                            str(current_config).strip() if current_config else None,
                            str(selection_config).strip() if selection_config else None,
                            str(rd_status).strip() if rd_status else None,
                        ]
                        has_meaningful = any(v and v not in NA_VALUES for v in excel_values)
                        if has_meaningful:
                            excel_pairs.add(pair_key)

                    key = (item.id, model.id)
                    if key in existing_values_map:
                        val = existing_values_map[key]
                        fields_to_update = [
                            ("final_config", final_config),
                            ("current_config", current_config),
                            ("selection_config", selection_config),
                            ("rd_status", rd_status),
                        ]
                        for field_name, excel_value in fields_to_update:
                            new_val = str(excel_value).strip() if excel_value else None
                            # 如果该 (IPN, 型号) 对在快照中不存在，则不创建"修改"草稿
                            # 该对会由下面的"新增"逻辑处理
                            if pair_key and pair_key not in snapshot_pairs:
                                # 注意：仍要更新ConfigValue（不跳过），仅跳过草稿创建
                                pass
                            # 快照中该配对4字段全为N/A：更新ConfigValue但不产生修改草稿（归类为"新增"）
                            skip_draft = (
                                (pair_key and pair_key not in snapshot_pairs) or
                                (pair_key and pair_key in snapshot_all_na_pairs)
                            )
                            # 从预计算快照值字典O(1)取值（替代原来的get_snapshot_value循环查找）
                            snap_val = snapshot_values.get((item_ipn, model.name, field_name))
                            old_val = getattr(val, field_name) if snap_val is None else snap_val
                            old_val_str = str(old_val).strip() if old_val else None
                            # 始终更新ConfigValue字段值（为新增/修改配对写入Excel数据）
                            if old_val_str != new_val:
                                setattr(val, field_name, new_val)
                            if skip_draft:
                                continue
                            if old_val_str != new_val:
                                draft_changes.append({
                                    "change_type": "update",
                                    "item_id": item.id,
                                    "model_id": model.id,
                                    "field_name": field_name,
                                    "old_value": old_val_str,
                                    "new_value": new_val,
                                    "rd_name": item_data.get('rd_name'),
                                })
                    elif key not in values_to_create:
                        # 创建新的配置值
                        value = ConfigValue(
                            item_id=item.id,
                            model_id=model.id,
                            final_config=str(final_config).strip() if final_config else None,
                            current_config=str(current_config).strip() if current_config else None,
                            selection_config=str(selection_config).strip() if selection_config else None,
                            rd_status=str(rd_status).strip() if rd_status else None
                        )
                        for field_name, excel_value in (
                            ("final_config", final_config),
                            ("current_config", current_config),
                            ("selection_config", selection_config),
                            ("rd_status", rd_status),
                        ):
                            new_val = str(excel_value).strip() if excel_value else None
                            # 如果该 (IPN, 型号) 对在快照中不存在，则不创建"修改"草稿
                            # 该对会由下面的"新增"逻辑处理
                            if pair_key and pair_key not in snapshot_pairs:
                                continue
                            # 快照中该配对4字段全为N/A，视为无数据，不产生修改草稿
                            if pair_key and pair_key in snapshot_all_na_pairs:
                                continue
                            # 从预计算快照值字典O(1)取值
                            snap_val = snapshot_values.get((item_ipn, model.name, field_name))
                            old_val_str = str(snap_val).strip() if snap_val else None
                            if old_val_str == new_val:
                                continue  # 快照值和Excel值相同，不创建草稿
                            draft_changes.append({
                                "change_type": "update",
                                "item_id": item.id,
                                "model_id": model.id,
                                "field_name": field_name,
                                "old_value": old_val_str,
                                "new_value": new_val,
                                "rd_name": item_data.get('rd_name'),
                            })
                        values_to_create[key] = value
                        values_created += 1

            # 批量添加新配置值
            for value in values_to_create.values():
                db.add(value)

            # 记录导入历史
            history = ImportHistory(
                series_id=series.id,
                filename=file.filename,
                records_count=items_created,
                status="success"
            )
            db.add(history)

            results.append({
                "series": current_series_name,
                "models": len(models),
                "items": items_created,
                "values": values_created,
                "changes": changes
            })
            change_log.extend(changes)

            # 构建 item_id → item 映射（用于后续查询）
            all_items_by_id = {}
            for item in existing_items_map.values():
                if item and item.id:
                    all_items_by_id[item.id] = item
            for item in processed_ipns.values():
                if item and item.id:
                    all_items_by_id[item.id] = item

            # 补充该系列所有型号到 models 列表和 excel_pairs
            # 确保后续变更检测覆盖该系列所有型号，而不是仅限于当前文件列范围
            series_models_from_db = await db.execute(
                select(ProductModel).where(ProductModel.series_id == series.id)
            )
            series_model_names = {m.name for m in models}
            for m in series_models_from_db.scalars().all():
                if m.name not in series_model_names:
                    models.append(m)
            # 注意：不补充已有ConfigValue配对到excel_pairs
            # excel_pairs 必须仅包含 Excel 中实际有意义的配对（含于第436-448行的NA过滤）
            # 否则全N/A的配对被误加入 excel_pairs 会变成"新增"

            # 创建草稿批次和草稿记录（用于前端展示变更）
            real_create_pairs = (excel_pairs - snapshot_pairs) | (excel_pairs & snapshot_all_na_pairs)
            real_delete_pairs = (snapshot_pairs - excel_pairs) - snapshot_all_na_pairs
            need_draft = bool(draft_changes) or bool(real_create_pairs) or bool(real_delete_pairs)
            if need_draft:
                draft_result = await db.execute(
                    select(DraftBatch).where(
                        DraftBatch.series_id == series.id,
                        DraftBatch.status == "draft"
                    ).order_by(DraftBatch.created_at.desc()).limit(1)
                )
                draft_batch = draft_result.scalar_one_or_none()

                if not draft_batch:
                    draft_batch = DraftBatch(
                        id=str(uuid.uuid4()),
                        series_id=series.id,
                        filename=file.filename,
                        status="draft",
                        total_count=0,
                        create_count=0,
                        update_count=0,
                        delete_count=0
                    )
                    db.add(draft_batch)
                    # 新增批次，无需清除草稿
                    batch_is_new = True
                else:
                    batch_is_new = False
                    # 确保计数不为 None（新创建的 DraftBatch 有 default=0，但从数据库加载的可能为 NULL）
                    if draft_batch.total_count is None:
                        draft_batch.total_count = 0
                    if draft_batch.create_count is None:
                        draft_batch.create_count = 0
                    if draft_batch.update_count is None:
                        draft_batch.update_count = 0
                    if draft_batch.delete_count is None:
                        draft_batch.delete_count = 0

                # 非新增批次：清除旧草稿，重新创建
                # 旧草稿可能包含之前（修复前）产生的错误数据
                existing_drafts_by_key = set()  # (change_type, item_id, model_id, field_name)
                if not batch_is_new:
                    # 清除该批次所有旧草稿，重新基于最新逻辑创建
                    old_drafts = await db.execute(
                        select(ConfigDraft).where(ConfigDraft.batch_id == draft_batch.id)
                    )
                    for d in old_drafts.scalars().all():
                        await db.delete(d)
                    # 重置计数（将重新计算）
                    draft_batch.total_count = 0
                    draft_batch.create_count = 0
                    draft_batch.update_count = 0
                    draft_batch.delete_count = 0

                # === 先筛选draft_changes：将"Excel全N/A但快照有有效值"的配对转为删除 ===
                # 必须在创建DB草稿之前完成，避免update草稿和delete草稿同时存在
                update_to_delete = set()  # (item_id, model_id) 转为删除的项
                converted_to_delete_pairs = set()  # (ipn, model_name)，避免 deleted_pairs 重复

                # 构建 item_id→ipn, model_id→name 映射
                item_id_to_ipn = {}
                for item in all_items_by_id.values():
                    if item and item.ipn:
                        item_id_to_ipn[item.id] = str(item.ipn).strip()
                model_id_to_name = {}
                for m in models:
                    if m.id:
                        model_id_to_name[m.id] = m.name

                # 按(item_id, model_id)分组
                groups = {}
                for entry in draft_changes:
                    key = (entry['item_id'], entry['model_id'])
                    if key not in groups:
                        groups[key] = []
                    groups[key].append(entry)

                filtered_changes = []
                for key, entries in groups.items():
                    item_id, model_id = key
                    ipn = item_id_to_ipn.get(item_id)
                    model_name = model_id_to_name.get(model_id)
                    pair_key = (ipn, model_name) if ipn and model_name else None

                    # 该配对应转为删除的判断：
                    # 快照中有有效值 且 Excel全为N/A（不在excel_pairs） 且 快照不全为N/A
                    should_be_delete = (
                        pair_key and
                        pair_key in snapshot_pairs and
                        pair_key not in excel_pairs and
                        pair_key not in snapshot_all_na_pairs
                    )
                    if should_be_delete:
                        update_to_delete.add(key)
                        converted_to_delete_pairs.add(pair_key)
                        continue
                    filtered_changes.extend(entries)

                # 计算被转换的条目数（用于修正count）
                converted_entries_count = sum(
                    len(entries) for key, entries in groups.items() if key in update_to_delete
                )
                draft_changes = filtered_changes

                # === 创建更新草稿（已排除应转为删除的配对） ===
                value_update_count = 0
                for entry in draft_changes:
                    draft_key = ("update", entry["item_id"], entry["model_id"], entry["field_name"])
                    if draft_key in existing_drafts_by_key:
                        continue
                    draft = ConfigDraft(
                        series_id=series.id,
                        batch_id=draft_batch.id,
                        change_type=entry["change_type"],
                        item_id=entry["item_id"],
                        model_id=entry["model_id"],
                        field_name=entry["field_name"],
                        old_value=entry["old_value"],
                        new_value=entry["new_value"]
                    )
                    db.add(draft)
                    value_update_count += 1

                draft_batch.update_count = value_update_count
                draft_batch.total_count = len(draft_changes) + len(update_to_delete)

                # 为"全部Excel值为N/A"的配对创建删除草稿
                for (del_item_id, del_model_id) in update_to_delete:
                    draft_key = ("delete", del_item_id, del_model_id, None)
                    if draft_key in existing_drafts_by_key:
                        continue
                    delete_draft = ConfigDraft(
                        series_id=series.id,
                        batch_id=draft_batch.id,
                        change_type="delete",
                        item_id=del_item_id,
                        model_id=del_model_id,
                        field_name=None,
                        old_value=None,
                        new_value=None
                    )
                    db.add(delete_draft)
                    draft_batch.delete_count += 1

                # 按机型创建"新增"草稿：每个 (IPN, 型号名) 对在 Excel 中有但快照中没有的
                # 或被快照标记为"全 N/A"（视为无数据）的配对
                models_by_name = {m.name: m for m in models}
                pairs_to_create = (excel_pairs - snapshot_pairs) | (excel_pairs & snapshot_all_na_pairs)
                for (pair_ipn, pair_model_name) in pairs_to_create:
                    # 查找 ConfigItem
                    pair_item = existing_items_map.get(pair_ipn) or processed_ipns.get(pair_ipn)
                    if not pair_item or not pair_item.id:
                        continue
                    # 查找 ProductModel
                    pair_model = models_by_name.get(pair_model_name)
                    if not pair_model or not pair_model.id:
                        continue
                    draft_key = ("create", pair_item.id, pair_model.id, None)
                    if draft_key in existing_drafts_by_key:
                        continue  # 已有相同新增草稿，跳过
                    create_draft = ConfigDraft(
                        series_id=series.id,
                        batch_id=draft_batch.id,
                        change_type="create",
                        item_id=pair_item.id,
                        model_id=pair_model.id,
                        field_name=None,
                        old_value=None,
                        new_value=pair_item.rd_name
                    )
                    db.add(create_draft)
                    draft_batch.create_count += 1
                    draft_batch.total_count += 1

                # 按机型创建"删除"草稿：每个 (IPN, 型号名) 对在快照中有但 Excel 中没有的
                # 排除快照中所有字段均为 N/A 的配对（视为无数据，不产生删除）
                deleted_pairs = (snapshot_pairs - excel_pairs) - snapshot_all_na_pairs - converted_to_delete_pairs
                if deleted_pairs:
                    # 批量查询所有可能需要的 ConfigItem 和 ProductModel
                    del_ipns = {p[0] for p in deleted_pairs}
                    del_model_names = {p[1] for p in deleted_pairs}
                    # 批量查 ConfigItem
                    del_items_by_ipn = {}
                    if del_ipns:
                        # 先收集内存中已有的
                        for ipn in del_ipns:
                            item = existing_items_map.get(ipn) or processed_ipns.get(ipn)
                            if item and item.id:
                                del_items_by_ipn[ipn] = item
                        # 再查库中遗漏的
                        missing_ipns = del_ipns - set(del_items_by_ipn.keys())
                        if missing_ipns:
                            del_items_result = await db.execute(
                                select(ConfigItem).where(ConfigItem.ipn.in_(missing_ipns))
                            )
                            for item in del_items_result.scalars().all():
                                if item.ipn:
                                    del_items_by_ipn[item.ipn] = item
                    # 批量查 ProductModel
                    del_models_by_name = {m.name: m for m in models}
                    missing_model_names = del_model_names - set(del_models_by_name.keys())
                    if missing_model_names:
                        del_models_result = await db.execute(
                            select(ProductModel).where(
                                ProductModel.series_id == series.id,
                                ProductModel.name.in_(missing_model_names)
                            )
                        )
                        for m in del_models_result.scalars().all():
                            del_models_by_name[m.name] = m

                    for (del_ipn, del_model_name) in deleted_pairs:
                        del_item = del_items_by_ipn.get(del_ipn)
                        del_model = del_models_by_name.get(del_model_name)
                        if not del_item or not del_item.id or not del_model or not del_model.id:
                            continue

                        # 检查该项在该机型下是否有有效值，跳过全 N/A 的垃圾数据
                        cv_result = await db.execute(
                            select(ConfigValue).where(
                                ConfigValue.item_id == del_item.id,
                                ConfigValue.model_id == del_model.id
                            )
                        )
                        cv = cv_result.scalar_one_or_none()
                        if cv:
                            has_meaningful = any(
                                getattr(cv, f, None) not in (None, '', 'N/A', '-')
                                for f in ('current_config', 'final_config', 'selection_config', 'rd_status')
                            )
                        else:
                            # ConfigValue 可能已被 clear_existing 删除，从快照判断
                            has_meaningful = any(
                                snapshot_values.get((del_ipn, del_model_name, f))
                                not in (None, '', 'N/A', '-')
                                for f in ('final_config', 'current_config', 'selection_config', 'rd_status')
                            )

                        if not has_meaningful:
                            # 全部 N/A，跳过（避免垃圾数据的删除草稿污染界面）
                            if del_item.ipn:
                                del_ipns.discard(del_ipn)
                            continue

                        delete_key = ("delete", del_item.id, del_model.id, None)
                        if delete_key in existing_drafts_by_key:
                            continue

                        delete_draft = ConfigDraft(
                            series_id=series.id,
                            batch_id=draft_batch.id,
                            change_type="delete",
                            item_id=del_item.id,
                            model_id=del_model.id,
                            field_name=None,
                            old_value=None,
                            new_value=None
                        )
                        db.add(delete_draft)
                        draft_batch.delete_count += 1
                        draft_batch.total_count += 1

        await db.commit()

    except HTTPException:
        # 重新抛出HTTP异常
        await db.rollback()
        raise
    except Exception as e:
        # 其他异常，回滚事务
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")

    return {
        "message": "导入成功",
        "details": results,
        "total_series": len(results),
        "change_log": change_log
    }


@router.post("/import/cleanup-duplicate-models")
async def cleanup_duplicate_models(db: AsyncSession = Depends(get_db)):
    """清理因系列范围合并Bug产生的重复机型归属

    旧的 merged_series 合并逻辑会扩大系列列范围，导致其他系列的机型
    被错误地创建到当前系列下。此端点检测同名机型跨系列重复的情况，
    自动保留在总型号数最少的系列中（正确系列），删除其他副本。
    """
    from app.models import ConfigValue, ConfigDraft

    # 1. 查找所有跨系列的同名机型
    # 按型号名称分组，筛选出出现在多个系列中的型号
    name_groups_result = await db.execute(
        select(
            ProductModel.name,
            func.count(ProductModel.id).label('total_instances'),
            func.count(func.distinct(ProductModel.series_id)).label('series_count')
        ).group_by(ProductModel.name)
        .having(func.count(func.distinct(ProductModel.series_id)) > 1)
    )
    duplicate_names_rows = name_groups_result.fetchall()
    duplicate_names = [r[0] for r in duplicate_names_rows]

    if not duplicate_names:
        return {"message": "没有发现重复的机型归属", "deleted": []}

    # 2. 获取所有重复机型及其所属系列
    # 先查询所有系列的总型号数索引
    all_models = await db.execute(select(ProductModel))
    all_models_list = all_models.scalars().all()

    # 统计每个系列的总型号数
    series_model_count = {}
    for m in all_models_list:
        series_model_count[m.series_id] = series_model_count.get(m.series_id, 0) + 1

    # 按名称分组
    models_by_name = {}
    for m in all_models_list:
        if m.name in duplicate_names:
            if m.name not in models_by_name:
                models_by_name[m.name] = []
            models_by_name[m.name].append(m)

    # 3. 对每组重复机型，决定保留和删除
    deleted_models = []  # [(model_id, model_name, series_name)]
    kept_models = []     # [(model_id, model_name, series_name)]

    for model_name, dup_models in models_by_name.items():
        # 按所属系列的总型号数升序排列
        dup_models_sorted = sorted(
            dup_models,
            key=lambda m: series_model_count.get(m.series_id, 0)
        )

        # 保留在总型号数最少的系列中的机型
        keep_model = dup_models_sorted[0]
        keep_series_result = await db.execute(
            select(ProductSeries).where(ProductSeries.id == keep_model.series_id)
        )
        keep_series = keep_series_result.scalar_one_or_none()
        kept_models.append((keep_model.id, model_name, keep_series.name if keep_series else "Unknown"))

        # 删除在其他系列中的机型
        for delete_model in dup_models_sorted[1:]:
            del_series_result = await db.execute(
                select(ProductSeries).where(ProductSeries.id == delete_model.series_id)
            )
            del_series = del_series_result.scalar_one_or_none()
            del_series_name = del_series.name if del_series else "Unknown"

            # 删除关联的配置值
            await db.execute(
                delete(ConfigValue).where(ConfigValue.model_id == delete_model.id)
            )
            # 删除关联的草稿
            await db.execute(
                delete(ConfigDraft).where(ConfigDraft.model_id == delete_model.id)
            )
            # 删除机型本身
            await db.delete(delete_model)

            deleted_models.append({
                "model_id": delete_model.id,
                "model_name": model_name,
                "deleted_from_series": del_series_name,
                "kept_in_series": keep_series.name if keep_series else "Unknown"
            })

    await db.commit()

    return {
        "message": f"清理完成，删除了 {len(deleted_models)} 个重复机型归属",
        "deleted": deleted_models,
        "kept": kept_models
    }


@router.post("/export")
async def export_excel(
    body: ExportRequest,
    db: AsyncSession = Depends(get_db)
):
    """导出Excel文件（所见即所得——由前端传入已筛选的行ID、机型ID和可见列字段）"""
    series_id = body.series_id
    include_main_unit = body.include_main_unit
    item_ids = body.item_ids
    categories = body.categories
    search = body.search
    model_ids = body.model_ids
    visible_fields = body.visible_fields
    draft_changes = body.draft_changes
    deleted_items = body.deleted_items
    new_items = body.new_items
    # 解析可见配置列
    ALL_CONFIG_FIELDS = ['final_config', 'current_config', 'selection_config', 'rd_status']
    FIELD_LABELS = {
        'final_config': '最终配置',
        'current_config': '当前配置',
        'selection_config': '选型类别',
        'rd_status': '研发状态',
    }
    if visible_fields:
        field_list = [f.strip() for f in visible_fields.split(',') if f.strip() in ALL_CONFIG_FIELDS]
    else:
        field_list = list(ALL_CONFIG_FIELDS)
    field_count = len(field_list)

    # 解析草稿变更记录
    draft_changes_map = {}
    if draft_changes:
        try:
            import json
            draft_changes_map = json.loads(draft_changes)
        except (json.JSONDecodeError, TypeError):
            pass

    # 解析删除项和新增项
    deleted_items_map = {}
    if deleted_items:
        try:
            import json
            deleted_items_map = json.loads(deleted_items)
        except (json.JSONDecodeError, TypeError):
            pass

    new_items_map = {}
    if new_items:
        try:
            import json
            new_items_map = json.loads(new_items)
        except (json.JSONDecodeError, TypeError):
            pass
    # 获取产品系列
    series_result = await db.execute(
        select(ProductSeries).where(ProductSeries.id == series_id)
    )
    series = series_result.scalar_one_or_none()

    if not series:
        raise HTTPException(status_code=404, detail="产品系列不存在")

    # 获取型号（支持按 model_ids 筛选）
    models_query = select(ProductModel).where(ProductModel.series_id == series_id)
    if model_ids:
        id_list = [int(x.strip()) for x in model_ids.split(',') if x.strip()]
        if id_list:
            models_query = models_query.where(ProductModel.id.in_(id_list))
    models_query = models_query.order_by(ProductModel.sort_order)
    models_result = await db.execute(models_query)
    models = models_result.scalars().all()

    if not models:
        raise HTTPException(status_code=400, detail="该系列下没有产品型号")

    # 获取配置项（按 item_ids 筛选——所见即所得；无 item_ids 时回退到 categories/search）
    items_query = select(ConfigItem).order_by(ConfigItem.row_index)
    if item_ids:
        id_list = [int(x.strip()) for x in item_ids.split(',') if x.strip()]
        if id_list:
            items_query = items_query.where(ConfigItem.id.in_(id_list))
    else:
        if categories:
            category_list = [c.strip() for c in categories.split(',') if c.strip()]
            if category_list:
                items_query = items_query.where(ConfigItem.category.in_(category_list))
        if search:
            items_query = items_query.where(
                (ConfigItem.rd_name.contains(search)) |
                (ConfigItem.ipn.contains(search)) |
                (ConfigItem.v_code.contains(search)) |
                (ConfigItem.zh_desc.contains(search)) |
                (ConfigItem.en_desc.contains(search))
            )
        if not include_main_unit:
            items_query = items_query.where(ConfigItem.category != "Main Unit")

    items_result = await db.execute(items_query)
    items = items_result.scalars().all()

    # 获取配置值
    model_id_list = [m.id for m in models]
    values_result = await db.execute(
        select(ConfigValue).where(ConfigValue.model_id.in_(model_id_list))
    )
    values = values_result.scalars().all()

    # 构建索引
    value_map = {}
    for v in values:
        if v.item_id not in value_map:
            value_map[v.item_id] = {}
        value_map[v.item_id][v.model_id] = v

    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "配置数据"

    # 定义样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    category_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # 草稿变更样式
    deleted_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    deleted_font = Font(strike=True, color="CC0000")
    new_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    updated_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # 第1行：产品系列
    ws.cell(row=1, column=1, value="研发名称")
    ws.cell(row=1, column=2, value="V代码")
    ws.cell(row=1, column=3, value="IPN号")
    ws.cell(row=1, column=4, value="中文描述")
    ws.cell(row=1, column=5, value="英文描述")

    col = 6
    for model in models:
        ws.cell(row=1, column=col, value=series.name)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + field_count - 1)
        col += field_count

    # 第2行：产品型号
    col = 6
    for model in models:
        ws.cell(row=2, column=col, value=model.name)
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + field_count - 1)
        col += field_count

    # 第3行：配置状态
    ws.cell(row=3, column=1, value="")  # A列
    ws.cell(row=3, column=2, value="")  # B列
    ws.cell(row=3, column=3, value="")  # C列
    ws.cell(row=3, column=4, value="")  # D列
    ws.cell(row=3, column=5, value="")  # E列

    col = 6
    for _ in models:
        for fi, field_name in enumerate(field_list):
            ws.cell(row=3, column=col + fi, value=FIELD_LABELS[field_name])
        col += field_count

    # 设置表头样式
    for row in range(1, 4):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    # 写入数据
    row_idx = 5
    current_category = None

    for item in items:
        # 如果分类变化，插入分类行
        if item.category and item.category != current_category:
            current_category = item.category
            ws.cell(row=row_idx, column=1, value=current_category)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ws.max_column)
            ws.cell(row=row_idx, column=1).fill = category_fill
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=1).alignment = center_align
            row_idx += 1

        # 写入固定列
        ws.cell(row=row_idx, column=1, value=item.rd_name)
        ws.cell(row=row_idx, column=2, value=item.v_code)
        ws.cell(row=row_idx, column=3, value=item.ipn)
        ws.cell(row=row_idx, column=4, value=item.zh_desc)
        ws.cell(row=row_idx, column=5, value=item.en_desc)

        # 写入配置值（仅可见列，含样式：绿=新增，红+删除线=删除，黄=修改）
        col = 6
        for model in models:
            model_val = value_map.get(item.id, {}).get(model.id)
            item_model_key = f"{item.id}_{model.id}"
            is_deleted = item_model_key in deleted_items_map
            is_new = item_model_key in new_items_map
            for fi, field_name in enumerate(field_list):
                cell = ws.cell(row=row_idx, column=col + fi)
                # 基础值：优先从删除快照取（因为删除了，DB值可能已不可靠）
                if is_deleted:
                    snap = deleted_items_map[item_model_key]
                    raw_value = snap.get(field_name, '') if isinstance(snap, dict) else ''
                else:
                    raw_value = getattr(model_val, field_name, '') if model_val else ''
                # 检查是否有逐格草稿变更
                change_key = f"{item.id}_{model.id}_{field_name}"
                has_update_change = False
                has_create_change = False
                if change_key in draft_changes_map:
                    dc = draft_changes_map[change_key]
                    if dc.get('changeType') == 'update' and dc.get('oldValue') is not None:
                        old_val = dc.get('oldValue', '') or '-'
                        new_val = dc.get('newValue', raw_value) or '-'
                        if old_val != '-':
                            raw_value = f"{old_val} → {new_val}"
                        else:
                            raw_value = new_val
                        has_update_change = True
                    elif dc.get('changeType') == 'create':
                        raw_value = dc.get('newValue', raw_value) or '-'
                        has_create_change = True
                # 应用样式
                display_value = raw_value or '-'
                if is_deleted:
                    cell.value = display_value
                    cell.fill = deleted_fill
                    cell.font = deleted_font
                elif has_update_change:
                    cell.value = display_value
                    cell.fill = updated_fill
                elif is_new or has_create_change:
                    cell.value = display_value
                    cell.fill = new_fill
                else:
                    cell.value = display_value
                cell.border = thin_border
            col += field_count

        # 设置边框
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col).border = thin_border

        row_idx += 1

    # 设置列宽
    ws.column_dimensions['A'].width = 35  # 研发名称
    ws.column_dimensions['B'].width = 12  # V代码
    ws.column_dimensions['C'].width = 12  # IPN号
    ws.column_dimensions['D'].width = 20  # 中文描述
    ws.column_dimensions['E'].width = 20  # 英文描述

    col = 6
    for _ in models:
        for fi in range(field_count):
            ws.column_dimensions[get_column_letter(col + fi)].width = 12
        col += field_count

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{series.name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.get("/template")
async def download_template():
    """下载导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "配置数据"

    # 表头
    headers = ["研发名称", "V代码", "IPN号", "中文描述", "英文描述"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 示例数据
    sample_data = [
        ["示例配置项1", "V001", "IPN-001", "中文描述1", "English Desc 1"],
        ["示例配置项2", "V002", "IPN-002", "中文描述2", "English Desc 2"],
    ]
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=config_template.xlsx"}
    )


@router.post("/preview")
async def preview_import(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """
    预览导入Excel文件（不实际导入）

    返回解析后的数据摘要，供用户确认后再导入
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持Excel文件(.xlsx, .xls)")

    # 读取文件
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # 解析合并单元格
    merged_info = parse_merged_cells(ws)

    # 解析产品系列（第1行）
    # 收集所有合并区域，按系列名合并
    series_ranges = {}
    processed_cols = set()

    for col in range(6, ws.max_column + 1):
        if col in processed_cols:
            continue
        if (1, col) in merged_info:
            info = merged_info[(1, col)]
            series_name = str(info['value']).strip() if info['value'] else None
            if series_name:
                for c in range(col, info['max_col'] + 1):
                    processed_cols.add(c)
                if series_name in series_ranges:
                    old_min, old_max = series_ranges[series_name]
                    series_ranges[series_name] = (min(old_min, col), max(old_max, info['max_col']))
                else:
                    series_ranges[series_name] = (col, info['max_col'])

    series_list = []
    for name, (col_start, col_end) in series_ranges.items():
        series_list.append({
            'name': name,
            'col_start': col_start,
            'col_end': col_end
        })

    if not series_list:
        series_list.append({
            'name': file.filename.replace('.xlsx', '').replace('.xls', ''),
            'col_start': 6,
            'col_end': ws.max_column
        })

    # 解析结果
    preview_result = {
        'filename': file.filename,
        'total_rows': ws.max_row,
        'total_cols': ws.max_column,
        'series': [],
        'summary': {
            'total_models': 0,
            'total_items': 0,
            'categories': []
        }
    }

    # 解析型号和数据
    for series_info in series_list:
        col_start = series_info['col_start']
        col_end = series_info['col_end']

        # 解析型号
        models = []
        col = col_start
        while col <= col_end:
            cell = ws.cell(row=2, column=col)
            if (2, col) in merged_info:
                info = merged_info[(2, col)]
                raw_value = info['value']
                model_end_col = info['max_col']
            else:
                raw_value = cell.value
                model_end_col = col + 3

            if raw_value:
                model_name = str(raw_value).split('//')[0].strip()
                models.append({
                    'name': model_name,
                    'col_start': col,
                    'col_end': model_end_col
                })
                col = model_end_col + 1
            else:
                col += 1

        # 解析配置项
        items = []
        categories = set()
        current_category = None

        for row_idx in range(4, ws.max_row + 1):
            first_col = ws.cell(row=row_idx, column=1).value

            if first_col and isinstance(first_col, str):
                stripped = first_col.strip()
                valid_categories = [
                    "Main Unit",
                    "Optional Features",
                    "Optional peripherals",
                    "*Optional peripherals(Preassemble in Factory)",
                    "Probes",
                    "Biopsy guide"
                ]
                if stripped in valid_categories or stripped.startswith("Optional"):
                    current_category = stripped
                    categories.add(stripped)
                    continue

            if current_category == "Main Unit":
                continue

            rd_name = ws.cell(row=row_idx, column=1).value
            ipn = ws.cell(row=row_idx, column=3).value

            if not rd_name and not ipn:
                continue

            items.append({
                'rd_name': str(rd_name).strip() if rd_name else None,
                'ipn': str(ipn).strip() if ipn else None,
                'category': current_category or 'Optional Features'
            })

        preview_result['series'].append({
            'name': series_info['name'],
            'models': [m['name'] for m in models],
            'item_count': len(items)
        })

        preview_result['summary']['total_models'] += len(models)
        preview_result['summary']['total_items'] += len(items)
        preview_result['summary']['categories'].extend(list(categories))

    preview_result['summary']['categories'] = list(set(preview_result['summary']['categories']))

    return preview_result