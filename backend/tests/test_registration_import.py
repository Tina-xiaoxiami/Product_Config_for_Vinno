import sqlite3

import openpyxl
from openpyxl import Workbook

from app.services import registration_rules
from app.services.registration_import import import_domestic_registration_workbook
from app.services.registration_migration import migrate_registration_schema
from app.services.registration_packages import stage_registration_package_draft
from app.services.registration_rules import parse_domestic_registration_workbook


def _create_database(path):
    connection = sqlite3.connect(path)
    connection.executescript(
        """
        PRAGMA foreign_keys = ON;
        CREATE TABLE knowledge_documents (
            id INTEGER PRIMARY KEY,
            document_type TEXT,
            title TEXT,
            file_name TEXT,
            file_path TEXT,
            sha256 TEXT,
            version TEXT,
            market TEXT,
            country TEXT,
            product_series TEXT,
            mime_type TEXT,
            source_status TEXT DEFAULT 'active'
        );
        INSERT INTO knowledge_documents VALUES (
            1, 'registration_difference', 'V10差异表', 'registration.xlsx',
            '/tmp/registration.xlsx', 'source-sha', '20250729',
            'domestic', 'CN', 'V10',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'active'
        );

        CREATE TABLE product_series (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL
        );
        INSERT INTO product_series VALUES
            (1, 'R&V10 series-China'),
            (2, 'R&V10 series-Oversea');

        CREATE TABLE product_models (
            id INTEGER PRIMARY KEY,
            series_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            status TEXT DEFAULT '生产中',
            column_start INTEGER,
            column_end INTEGER,
            sort_order INTEGER DEFAULT 0,
            config_group TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME,
            FOREIGN KEY(series_id) REFERENCES product_series(id)
        );
        INSERT INTO product_models (id, series_id, name, config_group) VALUES
            (1, 1, 'VINNO 10', NULL),
            (2, 1, 'VINNO 10E', NULL),
            (3, 1, 'VINNO 9', NULL),
            (4, 1, 'VINNO 9_Private', 'VINNO 9'),
            (5, 1, 'VINNO 9 综合版', NULL),
            (6, 2, 'VINNO 10', NULL);

        CREATE TABLE config_items (
            id INTEGER PRIMARY KEY,
            ipn TEXT,
            category TEXT,
            zh_desc TEXT
        );
        INSERT INTO config_items VALUES
            (10, '1000530', 'Probes', 'F2-5C探头'),
            (11, '1000744', 'Probes', 'G1-4P探头'),
            (12, '1000784', 'Probes', 'F4-9E探头');

        CREATE TABLE probe_models (
            id INTEGER PRIMARY KEY,
            model_number TEXT NOT NULL
        );
        CREATE TABLE probe_model_variants (
            id INTEGER PRIMARY KEY,
            probe_model_id INTEGER NOT NULL,
            internal_model TEXT NOT NULL,
            ipn TEXT,
            FOREIGN KEY(probe_model_id) REFERENCES probe_models(id)
        );
        INSERT INTO probe_models VALUES
            (21, 'F2-5C'), (22, 'G1-4P'), (23, 'F4-9E');
        INSERT INTO probe_model_variants VALUES
            (31, 21, 'F2-5C-Internal', '1000530'),
            (32, 22, 'G1-4P-Internal', '1000744'),
            (33, 23, 'F4-9E-Internal', '1000784');

        CREATE TABLE config_values (
            id INTEGER PRIMARY KEY,
            item_id INTEGER NOT NULL,
            model_id INTEGER NOT NULL,
            selection_config TEXT,
            current_config TEXT,
            FOREIGN KEY(item_id) REFERENCES config_items(id),
            FOREIGN KEY(model_id) REFERENCES product_models(id)
        );
        INSERT INTO config_values VALUES
            (1, 10, 1, 'X', 'X'),
            (2, 10, 2, '未定义', 'X'),
            (3, 11, 3, 'O', 'X'),
            (4, 11, 4, 'O', 'O'),
            (5, 11, 5, '未定义', 'Δ');
        """
    )
    connection.commit()
    connection.close()


def _write_registration_workbook(path, *, vinno10_unsupported="探头全适用"):
    workbook = Workbook()
    matrix = workbook.active
    matrix.title = "0729"
    matrix["A2"] = "支持探头\n共3把"
    matrix["B2"] = "F2-5C，G1-4P，F4-9E"
    matrix.append(["序号", "型号", "不支持探头", "通道数"])
    matrix.append([1, "VINNO 10", vinno10_unsupported, 128])
    matrix.append([2, "VINNO 10E", "F2-5C", 128])
    matrix.append([3, "VINNO 9", "F4-9E、G1-4P", 128])
    probes = workbook.create_sheet("Sheet1")
    probes.append([None, None, None])
    probes.append(["F2-5C", 1000530, 1000530])
    probes.append(["G1-4P", 1000744, 1000744])
    probes.append(["F4-9E", 1000784, 1000784])
    workbook.save(path)


def _write_registration_workbook_with_missing_ipns(path):
    workbook = Workbook()
    matrix = workbook.active
    matrix.title = "0729"
    matrix["A2"] = "支持探头\n共3把"
    matrix["B2"] = "S1-8CM，S1-8CX，SR1-10C"
    matrix.append(["序号", "型号", "不支持探头", "通道数"])
    matrix.append([1, "VINNO 10", "探头全适用", 128])
    probes = workbook.create_sheet("Sheet1")
    probes.append([None, None])
    probes.append(["S1-8CM", "1001335"])
    probes.append(["S1-8CX", None])
    probes.append(["SR1-10C", None])
    workbook.save(path)


class _UnsizedWorksheet:
    def __init__(self, worksheet):
        self._worksheet = worksheet
        self._sized = False

    @property
    def max_row(self):
        return self._worksheet.max_row if self._sized else None

    @property
    def max_column(self):
        return self._worksheet.max_column if self._sized else None

    def calculate_dimension(self, *, force=False):
        if force:
            self._sized = True
        return self._worksheet.calculate_dimension()

    def cell(self, *args, **kwargs):
        return self._worksheet.cell(*args, **kwargs)


class _UnsizedWorkbook:
    def __init__(self, workbook):
        self._workbook = workbook
        self.worksheets = [_UnsizedWorksheet(sheet) for sheet in workbook.worksheets]
        self.sheetnames = [sheet.title for sheet in workbook.worksheets]

    def __getitem__(self, name):
        return self.worksheets[self.sheetnames.index(name)]

    def close(self):
        self._workbook.close()


def test_parser_supports_valid_workbooks_without_dimension_metadata(tmp_path, monkeypatch):
    workbook_path = tmp_path / "registration.xlsx"
    _write_registration_workbook(workbook_path)
    actual = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
    unsized = _UnsizedWorkbook(actual)
    monkeypatch.setattr(registration_rules, "load_workbook", lambda *args, **kwargs: unsized)

    parsed = parse_domestic_registration_workbook(workbook_path)

    assert len(parsed.models) == 3
    assert len(parsed.probes) == 3


def test_registration_schema_and_import_materialize_country_model_probe_redlines(tmp_path):
    database_path = tmp_path / "product_config.db"
    workbook_path = tmp_path / "registration.xlsx"
    _create_database(database_path)
    _write_registration_workbook(workbook_path)

    migrate_registration_schema(database_path)
    first = import_domestic_registration_workbook(
        database_path,
        workbook_path,
        source_document_id=1,
    )
    repeated = import_domestic_registration_workbook(
        database_path,
        workbook_path,
        source_document_id=1,
    )
    _write_registration_workbook(workbook_path, vinno10_unsupported="F4-9E")
    changed = import_domestic_registration_workbook(
        database_path,
        workbook_path,
        source_document_id=1,
    )
    _write_registration_workbook(workbook_path)
    restored = import_domestic_registration_workbook(
        database_path,
        workbook_path,
        source_document_id=1,
    )

    assert first.model_count == 3
    assert first.probe_count == 3
    assert first.matrix_count == 9
    assert first.direct_link_count == 3
    assert first.derived_link_count == 2
    assert first.unmatched_product_model_count == 0
    assert first.new_snapshot is True
    assert repeated.new_snapshot is False
    assert changed.new_snapshot is True
    assert restored.new_snapshot is False

    connection = sqlite3.connect(database_path)
    assert connection.execute("PRAGMA integrity_check").fetchone()[0] == "ok"
    assert connection.execute("PRAGMA foreign_key_check").fetchall() == []
    assert connection.execute("SELECT COUNT(*) FROM registration_models").fetchone()[0] == 3
    assert connection.execute("SELECT COUNT(*) FROM registration_probes").fetchone()[0] == 3
    assert connection.execute("SELECT COUNT(*) FROM registration_model_probes").fetchone()[0] == 9
    assert connection.execute("SELECT COUNT(*) FROM registration_import_batches").fetchone()[0] == 2
    assert connection.execute(
        "SELECT COUNT(*) FROM registration_import_batches WHERE status = 'active'"
    ).fetchone()[0] == 1

    redlines = connection.execute(
        """
        SELECT model.model_name, probe.probe_model, matrix.registration_status
        FROM registration_model_probes matrix
        JOIN registration_models model ON model.id = matrix.registration_model_id
        JOIN registration_probes probe ON probe.id = matrix.registration_probe_id
        WHERE matrix.registration_status = 'unregistered'
        ORDER BY model.model_name, probe.probe_model
        """
    ).fetchall()
    assert redlines == [
        ("VINNO 10E", "F2-5C", "unregistered"),
        ("VINNO 9", "F4-9E", "unregistered"),
        ("VINNO 9", "G1-4P", "unregistered"),
    ]

    links = connection.execute(
        """
        SELECT product.name, registration.model_name, link.mapping_type
        FROM product_registration_model_links link
        JOIN product_models product ON product.id = link.product_model_id
        JOIN registration_models registration ON registration.id = link.registration_model_id
        ORDER BY product.id
        """
    ).fetchall()
    assert links == [
        ("VINNO 10", "VINNO 10", "direct"),
        ("VINNO 10E", "VINNO 10E", "direct"),
        ("VINNO 9", "VINNO 9", "direct"),
        ("VINNO 9_Private", "VINNO 9", "config_group"),
        ("VINNO 9 综合版", "VINNO 9", "confirmed_derived"),
    ]
    connection.close()


def test_registration_import_persists_multiple_probes_without_ipn(tmp_path):
    database_path = tmp_path / "product_config.db"
    workbook_path = tmp_path / "registration-without-ipns.xlsx"
    _create_database(database_path)
    _write_registration_workbook_with_missing_ipns(workbook_path)

    migrate_registration_schema(database_path)
    report = import_domestic_registration_workbook(
        database_path,
        workbook_path,
        source_document_id=1,
        confirmed_base_by_product_model_name={},
    )

    assert report.probe_count == 3
    connection = sqlite3.connect(database_path)
    assert connection.execute(
        "SELECT ipn FROM registration_probes WHERE probe_model = 'S1-8CX'"
    ).fetchone()[0] is None
    assert connection.execute(
        "SELECT ipn FROM registration_probes WHERE probe_model = 'SR1-10C'"
    ).fetchone()[0] is None
    assert connection.execute("PRAGMA foreign_key_check").fetchall() == []
    connection.close()


def test_registration_migration_preserves_legacy_probe_rows_when_ipn_becomes_optional(
    tmp_path,
):
    database_path = tmp_path / "legacy-product-config.db"
    _create_database(database_path)
    connection = sqlite3.connect(database_path)
    connection.executescript(
        """
        CREATE TABLE registration_import_batches (
            id INTEGER PRIMARY KEY,
            country_code TEXT NOT NULL,
            source_document_id INTEGER,
            source_version TEXT,
            source_sha256 TEXT,
            snapshot_hash TEXT NOT NULL,
            snapshot_json TEXT NOT NULL,
            model_count INTEGER NOT NULL,
            probe_count INTEGER NOT NULL,
            matrix_count INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'active',
            imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (country_code, snapshot_hash)
        );
        INSERT INTO registration_import_batches (
            id, country_code, snapshot_hash, snapshot_json,
            model_count, probe_count, matrix_count
        ) VALUES (1, 'CN', 'legacy', '{}', 0, 1, 0);
        CREATE TABLE registration_probes (
            id INTEGER PRIMARY KEY,
            country_code TEXT NOT NULL,
            probe_model TEXT NOT NULL,
            normalized_model TEXT NOT NULL,
            ipn TEXT NOT NULL,
            import_batch_id INTEGER NOT NULL REFERENCES registration_import_batches(id),
            source_document_id INTEGER,
            source_ref TEXT,
            source_status TEXT NOT NULL DEFAULT 'active',
            UNIQUE (country_code, normalized_model),
            UNIQUE (country_code, ipn)
        );
        INSERT INTO registration_probes (
            id, country_code, probe_model, normalized_model, ipn,
            import_batch_id, source_status
        ) VALUES (1, 'CN', 'F2-5C', 'f2-5c', '1000530', 1, 'active');
        """
    )
    connection.commit()
    connection.close()

    migrate_registration_schema(database_path)

    connection = sqlite3.connect(database_path)
    ipn_column = next(
        row for row in connection.execute("PRAGMA table_info(registration_probes)")
        if row[1] == "ipn"
    )
    assert ipn_column[3] == 0
    assert connection.execute(
        "SELECT probe_model, ipn FROM registration_probes"
    ).fetchall() == [("F2-5C", "1000530")]
    assert connection.execute("PRAGMA foreign_key_check").fetchall() == []
    connection.close()


def test_registration_migration_preserves_legacy_version_probe_rows(tmp_path):
    database_path = tmp_path / "legacy-version-probes.db"
    workbook_path = tmp_path / "registration.xlsx"
    certificate_path = tmp_path / "certificate.pdf"
    _create_database(database_path)
    _write_registration_workbook(workbook_path)
    certificate_path.write_bytes(b"%PDF-1.4 legacy version probe schema")
    migrate_registration_schema(database_path)
    draft = stage_registration_package_draft(
        database_path,
        country_code="CN",
        unit_code="LEGACY-VERSION-PROBES",
        display_name="旧版注册探头快照",
        product_series="V10",
        registration_number="TEST-LEGACY-VERSION-PROBES",
        certificate_path=certificate_path,
        difference_path=workbook_path,
        certificate_version="20260903",
        difference_version="20260903",
        confirmed_by="test",
        product_model_mappings={1: "VINNO 10"},
    )

    connection = sqlite3.connect(database_path)
    connection.execute("PRAGMA foreign_keys = OFF")
    connection.execute("DROP INDEX IF EXISTS ix_registration_version_probes_version")
    connection.execute("DROP INDEX IF EXISTS uq_registration_version_probe_ipn")
    connection.executescript(
        """
        CREATE TABLE registration_package_version_probes_legacy (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            version_id INTEGER NOT NULL
                REFERENCES registration_package_versions(id) ON DELETE CASCADE,
            registration_probe_id INTEGER NOT NULL
                REFERENCES registration_probes(id),
            probe_model TEXT NOT NULL,
            normalized_model TEXT NOT NULL,
            ipn TEXT NOT NULL,
            source_ref TEXT,
            UNIQUE (version_id, normalized_model),
            UNIQUE (version_id, ipn)
        );
        INSERT INTO registration_package_version_probes_legacy (
            id, version_id, registration_probe_id, probe_model,
            normalized_model, ipn, source_ref
        )
        SELECT id, version_id, registration_probe_id, probe_model,
               normalized_model, ipn, source_ref
        FROM registration_package_version_probes;
        DROP TABLE registration_package_version_probes;
        ALTER TABLE registration_package_version_probes_legacy
            RENAME TO registration_package_version_probes;
        CREATE INDEX ix_registration_version_probes_version
            ON registration_package_version_probes(version_id);
        """
    )
    connection.commit()
    connection.close()

    migrate_registration_schema(database_path)

    connection = sqlite3.connect(database_path)
    ipn_column = next(
        row
        for row in connection.execute(
            "PRAGMA table_info(registration_package_version_probes)"
        )
        if row[1] == "ipn"
    )
    assert ipn_column[3] == 0
    assert connection.execute(
        "SELECT COUNT(*) FROM registration_package_version_probes WHERE version_id = ?",
        (draft["id"],),
    ).fetchone()[0] == 3
    assert connection.execute("PRAGMA foreign_key_check").fetchall() == []
    connection.close()
