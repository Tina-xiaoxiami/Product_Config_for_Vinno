#!/usr/bin/env python3
"""
模拟导入逻辑，检查系列和机型的列范围映射
"""
import asyncio
import sys
sys.path.insert(0, '/Users/xiami/Documents/项目/产品配置管理系统/backend')

import openpyxl
from openpyxl.utils import get_column_letter

def parse_merged_cells(ws):
    """解析合并单元格"""
    merged_info = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
        value = ws.cell(row=min_row, column=min_col).value
        merged_info[(min_row, min_col)] = {
            'value': value,
            'max_col': max_col,
            'max_row': max_row
        }
    return merged_info

def analyze_excel_structure(filepath):
    """分析Excel结构"""
    print(f"\n分析文件: {filepath}")
    print("=" * 80)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    merged_info = parse_merged_cells(ws)

    # 解析系列（第1行，从F列开始）
    print("\n【系列解析】第1行:")
    series_ranges = {}
    processed_cols = set()

    for col in range(6, ws.max_column + 1):
        if col in processed_cols:
            continue
        if (1, col) in merged_info:
            info = merged_info[(1, col)]
            series_name = str(info['value']).strip() if info['value'] else None
            if series_name:
                for c in range(col, info['max_col'] + 1):
                    processed_cols.add(c)
                if series_name in series_ranges:
                    old_min, old_max = series_ranges[series_name]
                    series_ranges[series_name] = (min(old_min, col), max(old_max, info['max_col']))
                else:
                    series_ranges[series_name] = (col, info['max_col'])
                print(f"  {series_name}: 列 {get_column_letter(col)}-{get_column_letter(info['max_col'])} (col {col}-{info['max_col']})")

    # 解析型号（第2行）
    print("\n【型号解析】第2行:")
    for series_name, (col_start, col_end) in series_ranges.items():
        print(f"\n  系列 {series_name} (列 {col_start}-{col_end}):")
        col = col_start
        while col <= col_end:
            if (2, col) in merged_info:
                info = merged_info[(2, col)]
                model_name = str(info['value']).split('//')[0].strip() if info['value'] else None
                model_end_col = info['max_col']
            else:
                cell_value = ws.cell(row=2, column=col).value
                model_name = str(cell_value).split('//')[0].strip() if cell_value else None
                model_end_col = col + 3

            if model_name:
                print(f"    - {model_name}: 列 {get_column_letter(col)}-{get_column_letter(model_end_col)}")
                col = model_end_col + 1
            else:
                col += 1

    # 检查列范围是否有重叠
    print("\n【列范围重叠检查】:")
    series_list = [(name, range[0], range[1]) for name, range in series_ranges.items()]
    for i, (name1, start1, end1) in enumerate(series_list):
        for j, (name2, start2, end2) in enumerate(series_list):
            if i < j:
                # 检查是否有重叠
                if not (end1 < start2 or end2 < start1):
                    print(f"  ⚠️ {name1} 和 {name2} 列范围重叠!")
                    print(f"     {name1}: {start1}-{end1}")
                    print(f"     {name2}: {start2}-{end2}")

if __name__ == "__main__":
    # 检查上传目录中的Excel文件
    import os
    upload_dir = "/Users/xiami/Documents/项目/产品配置管理系统/backend/uploads"
    if os.path.exists(upload_dir):
        for filename in os.listdir(upload_dir):
            if filename.endswith(('.xlsx', '.xls')):
                filepath = os.path.join(upload_dir, filename)
                try:
                    analyze_excel_structure(filepath)
                except Exception as e:
                    print(f"分析 {filename} 失败: {e}")
    else:
        print(f"上传目录不存在: {upload_dir}")
